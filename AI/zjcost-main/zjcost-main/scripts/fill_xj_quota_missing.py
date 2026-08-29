from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


MISSING_ROWS = {
    "5-432": ("预制构件安装 过梁安装", "m3", "来源: 上册-203.png, 官方有此子目, 费用待核算"),
    "5-433": ("预制构件安装 小型构件安装(<=0.1m3)", "m3", "来源: 上册-203.png, 官方有此子目, 费用待核算"),
    "5-434": ("成品排气道安装 住宅厨卫排气道 适用楼层<=7", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-435": ("成品排气道安装 住宅厨卫排气道 适用楼层8~14", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-436": ("成品排气道安装 住宅厨卫排气道 适用楼层15~21", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-437": ("成品排气道安装 住宅厨卫排气道 适用楼层22~28", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-438": ("成品排气道安装 住宅厨卫排气道 适用楼层29~33", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-439": ("成品排气道安装 住宅相邻卫生间排气道安装(双孔) 适用楼层<=11", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-440": ("成品排气道安装 住宅相邻卫生间排气道安装(双孔) 适用楼层12~22", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-441": ("成品排气道安装 住宅相邻卫生间排气道安装(双孔) 适用楼层23~33", "节", "来源: 上册-204.png, 官方有此子目, 费用待核算"),
    "5-442": ("组合风帽安装", "套", "来源: 上册-205.png, 官方有此子目, 费用待核算"),
    "5-443": ("减隔震构件安装 叠层橡胶隔震支座安装 有效直径<=600mm", "个", "来源: 上册-205.png, 官方有此子目, 费用待核算"),
    "5-445": ("减隔震构件安装 叠层橡胶隔震支座安装 有效直径>900mm", "个", "来源: 上册-205.png, 官方有此子目, 费用待核算"),
    "5-446": ("减隔震构件安装 摩擦摆支座安装", "套", "来源: 上册-206.png, 官方有此子目, 费用待核算"),
    "5-448": ("阻尼器安装 双向式墙式支撑", "套", "来源: 上册-207.png, 官方有此子目, 费用待核算"),
    "5-449": ("阻尼器安装 单向式墙式支撑", "套", "来源: 上册-207.png, 官方有此子目, 费用待核算"),
    "5-450": ("阻尼器安装 单向式钢支撑", "套", "来源: 上册-207.png, 官方有此子目, 费用待核算"),
    "15-211": ("石材、瓷砖加工 石材倒角、磨边 <=10mm", "100m", "来源: 下册-230.png, 官方有此子目, 费用待核算"),
    "15-212": ("石材、瓷砖加工 石材倒角、磨边 >10mm", "100m", "来源: 下册-230.png, 官方有此子目, 费用待核算"),
    "15-213": ("石材、瓷砖加工 石材磨制、抛光 半圆边", "100m", "来源: 下册-230.png, 官方有此子目, 费用待核算"),
    "15-214": ("石材、瓷砖加工 石材磨制、抛光 加厚半圆边", "100m", "来源: 下册-230.png, 官方有此子目, 费用待核算"),
    "15-215": ("石材、瓷砖加工 石材开槽 断面面积<=30mm2", "100m", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-216": ("石材、瓷砖加工 石材开槽 断面面积<=100mm2", "100m", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-217": ("石材、瓷砖加工 石材开槽 断面面积<=200mm2", "100m", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-218": ("石材、瓷砖加工 石材开孔(周长)<=400mm", "100个", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-219": ("石材、瓷砖加工 石材开孔(周长)<=800mm", "100个", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-220": ("石材、瓷砖加工 石材开孔(周长)<=1000mm", "100个", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-221": ("石材、瓷砖加工 瓷砖倒角、抛光", "100m", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-222": ("石材、瓷砖加工 瓷砖开孔(周长)<=400mm", "100个", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-223": ("石材、瓷砖加工 瓷砖开孔(周长)<=800mm", "100个", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
    "15-224": ("石材、瓷砖加工 瓷砖开孔(周长)<=1000mm", "100个", "来源: 下册-231.png, 官方有此子目, 费用待核算"),
}


def code_key(code: str) -> tuple[int, ...]:
    return tuple(int(part) if part.isdigit() else 9999 for part in code.split("-"))


def find_source_workbook() -> Path:
    return next(path for path in (Path.home() / "Desktop").glob("*.xlsx") if "2020" in path.name)


def main() -> None:
    source = find_source_workbook()
    wb = load_workbook(source)
    ws = wb.active

    existing = {
        str(ws.cell(row, 1).value)
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 1).value
    }

    added = 0
    for code, (name, unit, remark) in sorted(MISSING_ROWS.items(), key=lambda item: code_key(item[0])):
        if code in existing:
            continue

        insert_at = ws.max_row + 1
        for row in range(2, ws.max_row + 1):
            current = ws.cell(row, 1).value
            if current and code_key(str(current)) > code_key(code):
                insert_at = row
                break

        ws.insert_rows(insert_at)
        ws.cell(insert_at, 1, code)
        ws.cell(insert_at, 2, name)
        ws.cell(insert_at, 3, unit)
        for col in range(4, 10):
            ws.cell(insert_at, col, None)
        ws.cell(insert_at, 10, None)
        ws.cell(insert_at, 11, remark)
        existing.add(code)
        added += 1

    out_dir = Path.cwd() / "output" / "xj_quota_check"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "xj_quota_filled_v3.xlsx"
    wb.save(out_path)
    print(f"added={added}")
    print(f"saved={out_path}")


if __name__ == "__main__":
    main()
