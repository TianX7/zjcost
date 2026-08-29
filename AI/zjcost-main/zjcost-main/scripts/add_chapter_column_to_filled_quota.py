from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


CHAPTERS = {
    1: "第1章 土石方工程",
    2: "第2章 地基处理与边坡支护工程",
    3: "第3章 桩基工程",
    4: "第4章 砌筑工程",
    5: "第5章 混凝土及钢筋混凝土工程",
    6: "第6章 金属结构工程",
    7: "第7章 木结构工程",
    8: "第8章 门窗工程",
    9: "第9章 屋面及防水工程",
    10: "第10章 保温隔热防腐工程",
    11: "第11章 楼地面装饰工程",
    12: "第12章 墙柱面装饰与隔断幕墙工程",
    13: "第13章 天棚工程",
    14: "第14章 油漆涂料裱糊工程",
    15: "第15章 其他装饰工程",
    16: "第16章 措施项目",
}


def infer_chapter(code: object) -> str:
    if code is None:
        return ""
    prefix = str(code).strip().split("-", 1)[0]
    if not prefix.isdigit():
        return ""
    return CHAPTERS.get(int(prefix), "")


def main() -> None:
    source = Path.cwd() / "output" / "xj_quota_check" / "xj_quota_filled_v3.xlsx"
    wb = load_workbook(source)
    ws = wb.active

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if "章节" in headers:
        chapter_col = headers.index("章节") + 1
    else:
        chapter_col = ws.max_column + 1
        ws.cell(1, chapter_col, "章节")

    for row in range(2, ws.max_row + 1):
        ws.cell(row, chapter_col, infer_chapter(ws.cell(row, 1).value))

    out = source.with_name("xj_quota_filled_v4_with_chapters.xlsx")
    wb.save(out)
    print(f"saved={out}")


if __name__ == "__main__":
    main()
