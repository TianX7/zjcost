from __future__ import annotations

import io
from collections import defaultdict

import openpyxl
from sqlalchemy.orm import Session

from app.models.quota_item import QuotaItem
from app.services.quota_import_service import (
    IMPORT_BATCH_SIZE,
    RETURN_ITEM_LIMIT,
    SQL_IN_CHUNK_SIZE,
    VALID_DISCIPLINES,
    QuotaImportStats,
    _chunks,
    _float_value,
    _normalize,
    infer_quota_chapter,
)

AUTO_DISCIPLINE = "AUTO"

_SHEET_DISCIPLINE_MAP: dict[str, str] = {
    "土建": "土建",
    "电气": "电气",
    "给排水": "给排水",
    "空调通风": "暖通消防",
    "工业管道": "暖通消防",
    "消防": "暖通消防",
    "仿古": "仿古",
    "古建": "仿古",
    "修缮": "仿古",
    "光伏": "光伏",
    "太阳能": "光伏",
    "组件": "光伏",
    "古渠": "水利灌溉",
    "灌溉": "水利灌溉",
    "水利": "水利灌溉",
    "渠道": "水利灌溉",
}


def _normalize_header_row(raw_headers: list[object] | tuple[object, ...]) -> dict[int, str]:
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(raw_headers):
        if cell is None:
            continue
        field = _normalize(str(cell))
        if field:
            col_map[idx] = field
    return col_map


def _normalize_sheet_name(sheet_name: str) -> str:
    return "".join(str(sheet_name).split())


def infer_discipline(
    sheet_name: str,
    quota_code: str = "",
    quota_name: str = "",
    requested_discipline: str = AUTO_DISCIPLINE,
) -> str:
    if requested_discipline and requested_discipline != AUTO_DISCIPLINE:
        return requested_discipline if requested_discipline in VALID_DISCIPLINES else "土建"

    normalized_sheet = _normalize_sheet_name(sheet_name)
    for key, value in _SHEET_DISCIPLINE_MAP.items():
        if key in normalized_sheet:
            return value

    text = f"{sheet_name} {quota_code} {quota_name}"
    if any(token in text for token in ("仿古", "古建", "古建筑", "斗拱", "青瓦", "筒瓦", "青砖", "修缮", "木作", "石作")):
        return "仿古"
    if any(token in text for token in ("光伏", "太阳能", "组件", "逆变器", "汇流箱", "组串", "并网", "PV1-F")):
        return "光伏"
    if any(token in text for token in ("古渠", "灌溉", "农渠", "斗渠", "渠道", "渠槽", "防渗", "闸门", "启闭机", "水利")):
        return "水利灌溉"
    if any(token in text for token in ("电气", "变压器", "电缆", "配电", "照明", "开关", "插座")):
        return "电气"
    if any(token in text for token in ("给排水", "给水", "排水", "喷淋", "管道")):
        return "给排水"
    if any(token in text for token in ("空调", "通风", "风管", "消防", "工业管道", "采暖", "暖通")):
        return "暖通消防"
    return "土建"


def inspect_quota_workbook(file_bytes: bytes) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheets: list[dict[str, object]] = []
        for index, ws in enumerate(wb.worksheets):
            row_iter = ws.iter_rows(values_only=True)
            raw_headers = next(row_iter, None)
            sample = next(row_iter, None)
            col_map = _normalize_header_row(raw_headers or [])
            matched_headers = sorted(set(col_map.values()))
            required = {"quota_code", "name", "unit"}
            quota_code = str(sample[0]).strip() if sample and len(sample) > 0 and sample[0] is not None else ""
            quota_name = str(sample[1]).strip() if sample and len(sample) > 1 and sample[1] is not None else ""
            sheets.append(
                {
                    "name": ws.title,
                    "index": index,
                    "rows": max((ws.max_row or 0) - 1, 0),
                    "columns": ws.max_column or 0,
                    "importable": required.issubset(set(col_map.values())),
                    "matched_headers": matched_headers,
                    "inferred_discipline": infer_discipline(ws.title, quota_code, quota_name),
                }
            )
        return sheets
    finally:
        wb.close()


def _selected_worksheets(
    wb: openpyxl.Workbook,
    sheet_names: list[str] | None,
) -> list[openpyxl.worksheet.worksheet.Worksheet]:
    if not sheet_names:
        return list(wb.worksheets)
    wanted = {str(name).strip() for name in sheet_names if str(name).strip()}
    if not wanted:
        return list(wb.worksheets)
    return [ws for ws in wb.worksheets if ws.title in wanted]


def parse_and_import_quota_with_sheets(
    file_bytes: bytes,
    db: Session,
    discipline: str = AUTO_DISCIPLINE,
    sheet_names: list[str] | None = None,
) -> QuotaImportStats:
    requested_discipline = discipline if discipline in VALID_DISCIPLINES else AUTO_DISCIPLINE
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        records_by_key: dict[tuple[str, str], dict[str, object]] = {}
        skipped = 0
        parsed_any_sheet = False

        for ws in _selected_worksheets(wb, sheet_names):
            row_iter = ws.iter_rows(values_only=True)
            raw_headers = next(row_iter, None)
            if raw_headers is None:
                continue

            col_map = _normalize_header_row(raw_headers)
            if not {"quota_code", "name", "unit"}.issubset(set(col_map.values())):
                continue

            parsed_any_sheet = True
            for row in row_iter:
                record: dict[str, object] = {}
                for idx, field in col_map.items():
                    val = row[idx] if idx < len(row) else None
                    if val is not None:
                        record[field] = val

                quota_code = str(record.get("quota_code", "")).strip()
                name = str(record.get("name", "")).strip()
                unit = str(record.get("unit", "")).strip()
                if not quota_code or not name or not unit:
                    skipped += 1
                    continue

                target_discipline = infer_discipline(
                    ws.title,
                    quota_code=quota_code,
                    quota_name=name,
                    requested_discipline=requested_discipline,
                )
                key = (target_discipline, quota_code)
                records_by_key[key] = {
                    "quota_code": quota_code,
                    "discipline": target_discipline,
                    "name": name,
                    "unit": unit,
                    "labor_qty": _float_value(record.get("labor_qty")),
                    "material_qty": _float_value(record.get("material_qty")),
                    "machine_qty": _float_value(record.get("machine_qty")),
                    "work_content": str(record.get("work_content", "")).strip(),
                    "applicable_scope": str(record.get("applicable_scope", "")).strip(),
                    "chapter": infer_quota_chapter(
                        quota_code,
                        discipline=target_discipline,
                        sheet_name=ws.title,
                        raw_chapter=record.get("chapter"),
                    ),
                    "version": str(record.get("version", "")).strip(),
                    "base_price": _float_value(record.get("base_price")),
                }

        if not parsed_any_sheet:
            return QuotaImportStats(imported=0, skipped=0, items=[])

        codes_by_discipline: dict[str, list[str]] = defaultdict(list)
        for discipline_name, quota_code in records_by_key:
            codes_by_discipline[discipline_name].append(quota_code)

        existing: dict[tuple[str, str], QuotaItem] = {}
        for discipline_name, codes in codes_by_discipline.items():
            for chunk in _chunks(codes, size=SQL_IN_CHUNK_SIZE):
                rows = (
                    db.query(QuotaItem)
                    .filter(QuotaItem.discipline == discipline_name, QuotaItem.quota_code.in_(chunk))
                    .all()
                )
                existing.update({(row.discipline, row.quota_code): row for row in rows})

        sample_items: list[QuotaItem] = []
        created = 0
        updated = 0
        for idx, (key, data) in enumerate(records_by_key.items(), 1):
            item = existing.get(key)
            if item is None:
                item = QuotaItem(**data)
                db.add(item)
                existing[key] = item
                created += 1
            else:
                for field, value in data.items():
                    if field != "quota_code":
                        setattr(item, field, value)
                updated += 1

            if len(sample_items) < RETURN_ITEM_LIMIT:
                sample_items.append(item)
            if idx % IMPORT_BATCH_SIZE == 0:
                db.flush()

        db.commit()
        for item in sample_items:
            db.refresh(item)

        return QuotaImportStats(
            imported=created + updated,
            skipped=skipped,
            items=sample_items,
            created=created,
            updated=updated,
        )
    except Exception:
        db.rollback()
        raise
    finally:
        wb.close()
