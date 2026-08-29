"""Services for importing quota workbooks and resource detail sheets."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

import openpyxl
from sqlalchemy.orm import Session

from app.models.quota_item import QuotaItem
from app.models.quota_resource_detail import QuotaResourceDetail
from app.models.quota_resource_material_mapping import QuotaResourceMaterialMapping

IMPORT_BATCH_SIZE = 1000
RETURN_ITEM_LIMIT = 200
SQL_IN_CHUNK_SIZE = 500
VALID_DISCIPLINES = {"土建", "给排水", "电气", "暖通消防", "仿古", "光伏", "水利灌溉", "旧材料"}

_HEADER_MAP: dict[str, str] = {
    "定额号": "quota_code",
    "定额编号": "quota_code",
    "定额编码": "quota_code",
    "子目编号": "quota_code",
    "编号": "quota_code",
    "quota_code": "quota_code",
    "code": "quota_code",
    "名称": "name",
    "定额名称": "name",
    "子目名称": "name",
    "项目名称": "name",
    "name": "name",
    "单位": "unit",
    "计量单位": "unit",
    "unit": "unit",
    "人工": "labor_qty",
    "人工含量": "labor_qty",
    "人工费": "labor_qty",
    "labor_qty": "labor_qty",
    "labor": "labor_qty",
    "材料": "material_qty",
    "材料含量": "material_qty",
    "材料费": "material_qty",
    "material_qty": "material_qty",
    "material": "material_qty",
    "机械": "machine_qty",
    "机械含量": "machine_qty",
    "机械费": "machine_qty",
    "machine_qty": "machine_qty",
    "machine": "machine_qty",
    "工作内容": "work_content",
    "项目特征": "work_content",
    "work_content": "work_content",
    "适用范围": "applicable_scope",
    "备注": "applicable_scope",
    "applicable_scope": "applicable_scope",
    "章节": "chapter",
    "章": "chapter",
    "分章": "chapter",
    "chapter": "chapter",
    "分部": "chapter",
    "版本": "version",
    "version": "version",
    "基价": "base_price",
    "综合单价": "base_price",
    "单价": "base_price",
    "base_price": "base_price",
    # ── 旧材料扩展字段（导入支持）──
    "获取方式": "acquisition_method",
    "acquisition_method": "acquisition_method",
    "来源说明": "origin_note",
    "origin_note": "origin_note",
    "遗址": "heritage_site",
    "遗址名称": "heritage_site",
    "文物名称": "heritage_site",
    "heritage_site": "heritage_site",
    "文物等级": "relic_level",
    "relic_level": "relic_level",
    "修复部位": "repair_part",
    "repair_part": "repair_part",
    "成色": "condition_grade",
    "成新率": "condition_grade",
    "condition_grade": "condition_grade",
    "批次号": "batch_no",
    "batch_no": "batch_no",
    "检测报告": "inspection_report_no",
    "检测报告编号": "inspection_report_no",
    "inspection_report_no": "inspection_report_no",
}

_BUILDING_CHAPTER_BY_PREFIX: dict[int, str] = {
    1: "第1章 土石方工程",
    2: "第2章 地基处理与边坡支护工程",
    3: "第3章 桩基工程",
    4: "第4章 砌筑工程",
    5: "第5章 混凝土及钢筋混凝土工程",
    6: "第6章 金属结构工程",
    7: "第7章 木结构工程",
    8: "第8章 门窗工程",
    9: "第9章 屋面及防水工程",
    10: "第10章 保温、隔热、防腐工程",
    11: "第11章 楼地面装饰工程",
    12: "第12章 墙柱面装饰与隔断幕墙工程",
    13: "第13章 天棚工程",
    14: "第14章 油漆涂料裱糊工程",
    15: "第15章 其他装饰工程",
    16: "第16章 措施项目",
}

_INSTALLATION_BOOK_BY_SHEET: dict[str, str] = {
    "电气": "第四册 电气设备安装工程",
    "给排水": "第十册 给排水、采暖、燃气工程",
    "空调通风": "第七册 通风空调工程",
    "工业管道": "第八册 工业管道安装工程",
    "消防": "第九册 消防工程",
    "仿古": "仿古建筑与古建修缮",
    "古建": "仿古建筑与古建修缮",
    "光伏": "光伏发电工程",
    "太阳能": "光伏发电工程",
    "灌溉": "水利灌溉工程",
    "古渠": "水利灌溉工程",
    "水利": "水利灌溉工程",
}

_INSTALLATION_BOOK_BY_DISCIPLINE: dict[str, str] = {
    "电气": "第四册 电气设备安装工程",
    "给排水": "第十册 给排水、采暖、燃气工程",
    "暖通消防": "暖通消防",
    "仿古": "仿古建筑与古建修缮",
    "光伏": "光伏发电工程",
    "水利灌溉": "水利灌溉工程",
    "旧材料": "遗址修复旧材料定额",
}

_RESOURCE_HEADER_MAP: dict[str, str] = {
    "定额号": "quota_code",
    "定额编号": "quota_code",
    "quota_code": "quota_code",
    "类别": "category",
    "category": "category",
    "资源类别": "category",
    "资源编码": "resource_code",
    "resource_code": "resource_code",
    "资源名称": "resource_name",
    "resource_name": "resource_name",
    "名称": "resource_name",
    "规格": "spec",
    "spec": "spec",
    "规格型号": "spec",
    "单位": "unit",
    "unit": "unit",
    "消耗量": "quantity",
    "quantity": "quantity",
    "含量": "quantity",
    "单价": "unit_price",
    "unit_price": "unit_price",
    "主材": "is_main_material",
    "is_main_material": "is_main_material",
}


@dataclass
class QuotaImportStats:
    imported: int
    skipped: int
    items: list[QuotaItem]
    created: int = 0
    updated: int = 0


@dataclass
class ResourceDetailImportStats:
    imported: int
    skipped: int
    quotas_updated: int
    replaced: int = 0


def _normalize(header: str) -> str | None:
    normalized = re.sub(r"[\s:：()（）\[\]【】]+", "", header.strip().lower())
    return _HEADER_MAP.get(normalized)


def _normalize_sheet_name(sheet_name: str) -> str:
    return "".join(str(sheet_name).split())


def _extract_numeric_prefix(quota_code: str) -> int | None:
    match = re.match(r"^(?:借)?(?:[A-Za-z])?\s*(\d{1,2})\s*[-－—]", quota_code.strip())
    if not match:
        return None
    return int(match.group(1))


def _infer_chapter(quota_code: str) -> str:
    prefix = _extract_numeric_prefix(quota_code)
    if prefix is None:
        return ""
    return _BUILDING_CHAPTER_BY_PREFIX.get(prefix, f"第{prefix}章")


def infer_quota_chapter(
    quota_code: str,
    discipline: str = "土建",
    sheet_name: str = "",
    raw_chapter: object | None = None,
) -> str:
    chapter = _normalize_chapter(raw_chapter)
    if chapter:
        return chapter

    normalized_sheet = _normalize_sheet_name(sheet_name)
    if normalized_sheet in _INSTALLATION_BOOK_BY_SHEET:
        return _INSTALLATION_BOOK_BY_SHEET[normalized_sheet]

    if discipline == "土建":
        return _infer_chapter(quota_code)

    return _INSTALLATION_BOOK_BY_DISCIPLINE.get(discipline, "")


def _chunks(values: list, size: int = SQL_IN_CHUNK_SIZE):
    for idx in range(0, len(values), size):
        yield values[idx:idx + size]


def _float_value(value: object) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def _normalize_chapter(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text in {"0", "0.0", "0.00", "-", "--", "无", "无章节", "nan", "None"}:
        return ""
    return text


def _normalize_acquisition_method(value: object) -> str:
    """将中文获取方式归一化为 recycle / reproduce / 空字符串。"""
    if value is None:
        return ""
    text = str(value).strip().lower()
    if not text:
        return ""
    if text in {"回收", "当地回收", "旧材料回收", "recycle", "回收旧材料"}:
        return "recycle"
    if text in {"复现", "原材料复现", "复刻", "仿制", "reproduce", "复现旧材料"}:
        return "reproduce"
    return text


def parse_and_import_quota(
    file_bytes: bytes,
    db: Session,
    discipline: str = "土建",
) -> QuotaImportStats:
    discipline = discipline if discipline in VALID_DISCIPLINES else "土建"
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        records_by_code: dict[str, dict[str, object]] = {}
        skipped = 0
        parsed_any_sheet = False

        for ws in wb.worksheets:
            row_iter = ws.iter_rows(values_only=True)
            raw_headers = next(row_iter, None)
            if raw_headers is None:
                continue

            col_map: dict[int, str] = {}
            for idx, cell in enumerate(raw_headers):
                if cell is None:
                    continue
                field = _normalize(str(cell))
                if field:
                    col_map[idx] = field

            if not {"quota_code", "name", "unit"}.issubset(set(col_map.values())):
                continue

            parsed_any_sheet = True
            for row in row_iter:
                record: dict[str, object] = {}
                for idx, field in col_map.items():
                    val = row[idx] if idx < len(row) else None
                    if val is not None:
                        record[field] = val

                quota_code = str(record.get("quota_code", "")).strip()
                name = str(record.get("name", "")).strip()
                unit = str(record.get("unit", "")).strip()
                if not quota_code or not name or not unit:
                    skipped += 1
                    continue

                records_by_code[quota_code] = {
                    "quota_code": quota_code,
                    "discipline": discipline,
                    "name": name,
                    "unit": unit,
                    "labor_qty": _float_value(record.get("labor_qty")),
                    "material_qty": _float_value(record.get("material_qty")),
                    "machine_qty": _float_value(record.get("machine_qty")),
                    "work_content": str(record.get("work_content", "")).strip(),
                    "applicable_scope": str(record.get("applicable_scope", "")).strip(),
                    "chapter": infer_quota_chapter(
                        quota_code,
                        discipline=discipline,
                        sheet_name=ws.title,
                        raw_chapter=record.get("chapter"),
                    ),
                    "version": str(record.get("version", "")).strip(),
                    "base_price": _float_value(record.get("base_price")),
                    # ── 旧材料扩展字段 ──
                    "acquisition_method": _normalize_acquisition_method(
                        record.get("acquisition_method")
                    ),
                    "origin_note": str(record.get("origin_note", "")).strip(),
                    "heritage_site": str(record.get("heritage_site", "")).strip(),
                    "relic_level": str(record.get("relic_level", "")).strip(),
                    "repair_part": str(record.get("repair_part", "")).strip(),
                    "condition_grade": str(record.get("condition_grade", "")).strip(),
                    "batch_no": str(record.get("batch_no", "")).strip(),
                    "inspection_report_no": str(record.get("inspection_report_no", "")).strip(),
                }

        if not parsed_any_sheet:
            return QuotaImportStats(imported=0, skipped=0, items=[])

        codes = list(records_by_code)
        existing: dict[str, QuotaItem] = {}
        for chunk in _chunks(codes):
            rows = (
                db.query(QuotaItem)
                .filter(QuotaItem.discipline == discipline, QuotaItem.quota_code.in_(chunk))
                .all()
            )
            existing.update({row.quota_code: row for row in rows})

        sample_items: list[QuotaItem] = []
        created = 0
        updated = 0
        for idx, (quota_code, data) in enumerate(records_by_code.items(), 1):
            item = existing.get(quota_code)
            if item is None:
                item = QuotaItem(**data)
                db.add(item)
                existing[quota_code] = item
                created += 1
            else:
                for key, value in data.items():
                    if key != "quota_code":
                        setattr(item, key, value)
                updated += 1

            if len(sample_items) < RETURN_ITEM_LIMIT:
                sample_items.append(item)
            if idx % IMPORT_BATCH_SIZE == 0:
                db.flush()

        db.commit()
        for item in sample_items:
            db.refresh(item)

        return QuotaImportStats(
            imported=created + updated,
            skipped=skipped,
            items=sample_items,
            created=created,
            updated=updated,
        )
    except Exception:
        db.rollback()
        raise
    finally:
        wb.close()


def parse_and_import_resource_details(
    file_bytes: bytes,
    db: Session,
) -> ResourceDetailImportStats:
    """Import quota resource details from Excel."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return ResourceDetailImportStats(imported=0, skipped=0, quotas_updated=0)

        row_iter = ws.iter_rows(values_only=True)
        raw_headers = next(row_iter, None)
        if raw_headers is None:
            return ResourceDetailImportStats(imported=0, skipped=0, quotas_updated=0)

        col_map: dict[int, str] = {}
        for idx, cell in enumerate(raw_headers):
            if cell is None:
                continue
            field = _RESOURCE_HEADER_MAP.get(str(cell).strip().lower())
            if field:
                col_map[idx] = field

        if not {"quota_code", "category", "resource_name", "unit"}.issubset(set(col_map.values())):
            skipped = max((ws.max_row or 1) - 1, 0)
            return ResourceDetailImportStats(imported=0, skipped=skipped, quotas_updated=0)

        records: list[dict[str, object]] = []
        quota_codes: set[str] = set()
        skipped = 0

        for row in row_iter:
            record: dict[str, object] = {}
            for idx, field in col_map.items():
                val = row[idx] if idx < len(row) else None
                if val is not None:
                    record[field] = val

            quota_code = str(record.get("quota_code", "")).strip()
            category = str(record.get("category", "")).strip()
            resource_name = str(record.get("resource_name", "")).strip()
            unit = str(record.get("unit", "")).strip()

            if not quota_code or not category or not resource_name or not unit:
                skipped += 1
                continue
            if category not in ("人工", "材料", "机械"):
                skipped += 1
                continue

            is_main = str(record.get("is_main_material", "")).strip().lower()
            records.append(
                {
                    "quota_code": quota_code,
                    "category": category,
                    "resource_code": str(record.get("resource_code", "")).strip(),
                    "resource_name": resource_name,
                    "spec": str(record.get("spec", "")).strip(),
                    "unit": unit,
                    "quantity": _float_value(record.get("quantity")),
                    "unit_price": _float_value(record.get("unit_price")),
                    "is_main_material": 1 if is_main in ("1", "是", "yes", "true", "y") else 0,
                }
            )
            quota_codes.add(quota_code)

        quota_lookup: dict[str, QuotaItem] = {}
        for chunk in _chunks(list(quota_codes)):
            rows = db.query(QuotaItem).filter(QuotaItem.quota_code.in_(chunk)).all()
            quota_lookup.update({row.quota_code: row for row in rows})

        valid_records: list[tuple[QuotaItem, dict[str, object]]] = []
        updated_quota_ids: set[int] = set()
        for record in records:
            quota = quota_lookup.get(str(record["quota_code"]))
            if not quota:
                skipped += 1
                continue
            valid_records.append((quota, record))
            updated_quota_ids.add(quota.id)

        replaced = 0
        quota_id_list = list(updated_quota_ids)
        for quota_chunk in _chunks(quota_id_list):
            detail_ids = [
                row[0]
                for row in db.query(QuotaResourceDetail.id)
                .filter(QuotaResourceDetail.quota_item_id.in_(quota_chunk))
                .all()
            ]
            for detail_chunk in _chunks(detail_ids):
                db.query(QuotaResourceMaterialMapping).filter(
                    QuotaResourceMaterialMapping.resource_detail_id.in_(detail_chunk)
                ).delete(synchronize_session=False)
            if detail_ids:
                replaced += db.query(QuotaResourceDetail).filter(
                    QuotaResourceDetail.quota_item_id.in_(quota_chunk)
                ).delete(synchronize_session=False)

        imported = 0
        for idx, (quota, record) in enumerate(valid_records, 1):
            db.add(
                QuotaResourceDetail(
                    quota_item_id=quota.id,
                    category=str(record["category"]),
                    resource_code=str(record.get("resource_code", "")),
                    resource_name=str(record["resource_name"]),
                    spec=str(record.get("spec", "")),
                    unit=str(record["unit"]),
                    quantity=float(record.get("quantity", 0)),
                    unit_price=float(record.get("unit_price", 0)),
                    is_main_material=int(record.get("is_main_material", 0)),
                )
            )
            imported += 1
            if idx % IMPORT_BATCH_SIZE == 0:
                db.flush()

        for quota_chunk in _chunks(quota_id_list):
            db.query(QuotaItem).filter(QuotaItem.id.in_(quota_chunk)).update(
                {QuotaItem.has_resource_details: 1},
                synchronize_session=False,
            )

        db.commit()
        return ResourceDetailImportStats(
            imported=imported,
            skipped=skipped,
            quotas_updated=len(updated_quota_ids),
            replaced=replaced,
        )
    except Exception:
        db.rollback()
        raise
    finally:
        wb.close()
