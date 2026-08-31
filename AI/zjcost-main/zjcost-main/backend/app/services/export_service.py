"""Export service: generate Excel valuation reports."""

from __future__ import annotations

import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from app.models.boq_item import BoqItem
from app.models.calc_result import CalcResult
from app.models.line_item_quota_binding import LineItemQuotaBinding
from app.models.project import Project
from app.models.quota_item import QuotaItem
from app.services.pricing_engine import DEFAULT_FEE_CONFIG
from app.services.project_calc_service import run_project_calculation
from app.services.snapshot_service import diff_snapshots


_ORDER_KEYWORDS: list[tuple[str, int]] = [
    ("土石方", 0), ("平整场地", 0),
    ("地基", 1), ("基础", 1), ("桩", 1),
    ("混凝土", 2), ("钢筋", 2), ("砌筑", 2), ("砌体", 2), ("主体", 2),
    ("屋面", 3), ("防水", 3), ("保温", 3), ("隔热", 3),
    ("楼地面", 4), ("墙柱面", 4), ("天棚", 4), ("吊顶", 4), ("装饰", 4), ("装修", 4), ("油漆", 4), ("涂料", 4),
    ("门窗", 5),
    ("给排水", 7), ("消防", 7), ("通风", 7), ("空调", 7), ("暖通", 7), ("水暖", 7),
    ("电气", 8), ("弱电", 8), ("智能化", 8),
    ("措施", 9), ("模板", 9), ("脚手架", 9), ("临时", 9),
]


def division_rank(division: str) -> int:
    """按工程施工顺序给分部排序：基础→主体→屋面防水→装饰装修→门窗→安装→电气→措施，未识别分部沉底。"""
    dd = division or ""
    best = 99
    for kw, rank in _ORDER_KEYWORDS:
        if kw in dd and rank < best:
            best = rank
    return best


def _cn_upper(amount: float) -> str:
    """数字金额转中文大写（亿/万/个三段式，演示精度到分）。"""
    digits = "零壹贰叁肆伍陆柒捌玖"
    units = ["", "拾", "佰", "仟"]
    if not amount or amount <= 0:
        return "零元整"
    amt = int(round(float(amount) * 100))
    yuan, jiao, fen = amt // 100, (amt % 100) // 10, amt % 10

    def seg(n: int) -> str:
        out, zero = "", False
        for i in range(3, -1, -1):
            d = (n // (10 ** i)) % 10
            if d == 0:
                if out and out[-1] != "零":
                    zero = True
            else:
                if zero:
                    out += "零"
                    zero = False
                out += digits[d] + units[i]
        return out

    out = ""
    if yuan >= 100000000:
        out += seg(yuan // 100000000) + "亿"
    wan = (yuan % 100000000) // 10000
    if wan:
        out += seg(wan) + "万"
    ge = yuan % 10000
    if ge:
        if out and ge < 1000 and not seg(ge).startswith("零"):
            out += "零"
        out += seg(ge)
    if not out:
        out = "零"
    out += "元"
    if jiao == 0 and fen == 0:
        return out + "整"
    if jiao:
        out += digits[jiao] + "角"
    if fen:
        out += digits[fen] + "分"
    return out


def _sx_border(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c <= 1 else "left", vertical="center", wrap_text=(c in (3, 4))
            )


def _sx_title(ws, row: int, end_col: int, text: str, size: int = 16) -> None:
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=size)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")


def _sx_header(ws, row: int, headers: list[str], fill: str = "D9EAF7") -> None:
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _sx_widths(ws, widths: list[int]) -> None:
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w


def export_valuation_report(project_id: int, db: Session) -> bytes:
    """按《实训手册》版式生成招标控制价 Excel（14 表）。"""

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise ValueError(f"Project {project_id} not found")

    summary, line_results = run_project_calculation(project_id=project_id, db=db)

    grand = float(summary.grand_total or 0)
    total_boq = float(sum((r.total or 0) for _, r in line_results))
    regulatory = float(summary.total_regulatory or 0)
    tax = float(summary.total_tax or 0)
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    project_name = project.name or "未命名项目"
    bid_section = f"{project.region} {project_name}工程"
    rows = sorted(line_results, key=lambda x: (division_rank(x[0].division or ""), x[0].sort_order))

    div_sums: dict[str, float] = {}
    for boq, result in rows:
        d = boq.division or "未分类"
        div_sums[d] = div_sums.get(d, 0.0) + float(result.total or 0)

    thin = Side(style="thin", color="000000")
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    R = Alignment(horizontal="right", vertical="center")
    HDR = Font(bold=True, size=10)

    def put(ws, addr, value, font=None, align=None, border=True):
        cell = ws[addr]
        cell.value = value
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if border:
            cell.border = BD
        return cell

    def box(ws, r1, r2, c1, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = BD

    def title_block(ws, title, last_col_letter, page="第  1  页  共  1  页"):
        ws.merge_cells(f"A2:{last_col_letter}2")
        t = ws["A2"]
        t.value = title
        t.font = Font(bold=True, size=14)
        t.alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "C3", f"标段：{bid_section}", align=L, border=False)

    wb = openpyxl.Workbook()
    try:
        # ================= 封-2 =================
        ws = wb.active
        ws.title = "封-2 招标控制价封面"
        ws.merge_cells("A1:B1"); ws.merge_cells("F1:G1")
        put(ws, "F1", "工程", border=False)
        ws.merge_cells("A2:G2")
        a2 = ws["A2"]; a2.value = "招标控制价"; a2.font = Font(bold=True, size=26); a2.alignment = C
        ws.merge_cells("A3:G3")
        put(ws, "A3", f"（{project_name}）", align=C, border=False)
        ws.merge_cells("B4:C4"); put(ws, "B4", "招  标  人：", align=L)
        ws.merge_cells("D4:F4"); put(ws, "D4", "＿＿＿＿＿＿＿＿＿＿＿＿", align=L)
        ws.merge_cells("B5:C5"); put(ws, "B5", "招标控制价（小写）：", align=L)
        ws.merge_cells("D5:F5"); put(ws, "D5", f"{grand:,.2f} 元", align=L)
        put(ws, "B6", "（大写）：", align=L)
        ws.merge_cells("D6:E6"); put(ws, "D6", _cn_upper(grand), align=L)
        put(ws, "B7", "编制时间：", align=L, border=False)
        put(ws, "D7", generated, align=L, border=False)
        put(ws, "D8", "", border=False); ws.merge_cells("D8:E8"); put(ws, "D8", "封-2", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFG", [12, 12, 14, 14, 14, 14, 14]):
            ws.column_dimensions[col].width = w

        # ================= 扉-2 =================
        ws = wb.create_sheet("扉-2 招标控制价扉页")
        ws.merge_cells("A1:B1"); ws.merge_cells("C1:G1")
        put(ws, "H1", "工程", border=False)
        ws.merge_cells("C2:G2")
        c2 = ws["C2"]; c2.value = "招标控制价"; c2.font = Font(bold=True, size=22); c2.alignment = C
        put(ws, "A3", "招标控制价", align=L)
        ws.merge_cells("B3:C3"); put(ws, "B3", "（小写）：", align=L)
        ws.merge_cells("D3:H3"); put(ws, "D3", f"{grand:,.2f} 元", align=L)
        ws.merge_cells("B4:C4"); put(ws, "B4", "（大写）：", align=L)
        ws.merge_cells("D4:H4"); put(ws, "D4", _cn_upper(grand), align=L)
        ws.merge_cells("B5:C5"); put(ws, "B5", "招  标  人：", align=L)
        ws.merge_cells("G5:H5"); put(ws, "G5", "＿＿＿＿＿＿", align=L)
        ws.merge_cells("A6:B6"); put(ws, "A6", "编制时间：", align=L, border=False)
        ws.merge_cells("D6:E6"); put(ws, "D6", generated, align=L, border=False)
        put(ws, "H6", "扉—2", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFGH", [12, 12, 12, 12, 12, 12, 12, 10]):
            ws.column_dimensions[col].width = w

        # ================= 表-01 总说明 =================
        ws = wb.create_sheet("表-01 总说明")
        ws.merge_cells("B1:D1")
        b1 = ws["B1"]; b1.value = "总  说  明"; b1.font = Font(bold=True, size=16); b1.alignment = C
        put(ws, "A2", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "E2", "第  1  页  共  1  页", align=C, border=False)
        ws.merge_cells("A3:E4")
        notes = (
            f"一、工程概况：本工程为{project_name}，位于{project.region}；计价标准 {project.standard_type}；币种 {project.currency}。\n"
            f"二、招标控制价合计 {grand:,.2f} 元，其中分部分项工程费 {total_boq:,.2f} 元，规费 {regulatory:,.2f} 元，税金 {tax:,.2f} 元。\n"
            "三、综合单价包含人工费、材料费、机械费、管理费、利润、规费和税金。\n"
            "四、措施项目费、其他项目费本工程暂按 0 填报，若发生按实调整。"
        )
        a3 = ws["A3"]; a3.value = notes; a3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        box(ws, 3, 4, 1, 5)
        put(ws, "E5", "表-01", border=False).font = Font(size=9)
        for col, w in zip("ABCDE", [16, 22, 22, 22, 18]):
            ws.column_dimensions[col].width = w

        # ================= 表-04 单位工程招标控制价汇总表 =================
        ws = wb.create_sheet("表-04 单位工程招标控制价汇总表")
        title_block(ws, "单位工程招标控制价汇总表", "F")
        put(ws, "E3", "第  1  页  共  1  页", align=C, border=False)
        put(ws, "A4", "序号", font=HDR, align=C)
        put(ws, "B4", "汇总内容", font=HDR, align=C)
        put(ws, "D4", "金额（元）", font=HDR, align=C)
        put(ws, "F4", "其中：暂估价(元)", font=HDR, align=C)
        r = 5
        def sum_row(no, name, val):
            nonlocal r
            ws.merge_cells(f"B{r}:C{r}")
            put(ws, f"A{r}", no, align=C); put(ws, f"B{r}", name, align=L)
            ws.merge_cells(f"D{r}:E{r}"); put(ws, f"D{r}", f"{val:,.2f}" if val else "", align=R)
            put(ws, f"F{r}", "－" if val is None else "", align=C)
            r += 1
        sum_row("一", "分部分项工程费", total_boq)
        for i, (d, s) in enumerate(sorted(div_sums.items(), key=lambda kv: division_rank(kv[0])), 1):
            sum_row(f"1.{i}", d, s)
        sum_row("二", "措施项目费", 0.0)
        sum_row("2.1", "单价措施项目费", 0.0)
        sum_row("2.2", "总价措施项目费", 0.0)
        sum_row("2.21", "其中：安全文明施工费", 0.0)
        sum_row("三", "其他项目费", 0.0)
        sum_row("3.1", "暂列金额", 0.0)
        sum_row("3.2", "暂估价", 0.0)
        sum_row("3.3", "计日工", 0.0)
        sum_row("3.4", "总承包服务费", 0.0)
        sum_row("四", "规费", regulatory)
        sum_row("五", "税前工程造价", total_boq + regulatory)
        sum_row("六", "税金", tax)
        put(ws, f"A{r + 1}", "招标控制价合计=一+二+三+四+六", align=L, border=False)
        put(ws, f"A{r + 2}", "注：本表适用于单位工程招标控制价或投标报价的汇总，如无单位工程划分，单项工程也使用本表汇总", align=L, border=False)
        put(ws, f"E{r + 3}", "表—04", border=False).font = Font(size=9)
        for col, w in zip("ABCDEF", [8, 16, 16, 14, 14, 18]):
            ws.column_dimensions[col].width = w

        # ================= 表-08 分部分项工程和单价措施项目清单与计价表 =================
        ws = wb.create_sheet("表-08 分部分项工程清单与计价表")
        put(ws, "A2", "分部分项工程和单价措施项目清单与计价表", border=False)
        ws.merge_cells("A2:I2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "E3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "I3", "第  1  页  共  1  页", align=C, border=False)
        for addr, v in [("A4", "序号"), ("B4", "项目编码"), ("C4", "项目名称"),
                        ("D4", "项目特征描述"), ("E4", "计量单位"), ("F4", "工程量")]:
            put(ws, addr, v, font=HDR, align=C)
        ws.merge_cells("G4:I4"); put(ws, "G4", "金额（元）", font=HDR, align=C)
        put(ws, "G5", "综合单价", font=HDR, align=C)
        put(ws, "H5", "合价", font=HDR, align=C)
        put(ws, "I5", "其中", font=HDR, align=C)
        put(ws, "I6", "暂估价", font=HDR, align=C)
        for a in ("A", "B", "C", "E", "F"):
            ws.merge_cells(f"{a}4:{a}6")
        ws.merge_cells("D4:D6")
        ws.merge_cells("G5:G6"); ws.merge_cells("H5:H6")
        rr = 7
        for i, (boq, result) in enumerate(rows, 1):
            qty = float(boq.quantity or 1) or 1
            total = float(result.total or 0)
            put(ws, f"A{rr}", i, align=C)
            put(ws, f"B{rr}", boq.code or "", align=L)
            put(ws, f"C{rr}", boq.name or "", align=L)
            put(ws, f"D{rr}", boq.division or "未分类", align=L)
            put(ws, f"E{rr}", boq.unit or "", align=C)
            put(ws, f"F{rr}", round(qty, 2), align=R)
            put(ws, f"G{rr}", round(total / qty, 2), align=R)
            put(ws, f"H{rr}", round(total, 2), align=R)
            put(ws, f"I{rr}", 0, align=R)
            rr += 1
        if not rows:
            put(ws, "C7", "（无清单数据）", align=L); rr = 8
        put(ws, f"A{rr}", "本页小计", align=L, border=False)
        put(ws, f"H{rr}", round(total_boq, 2), align=R)
        put(ws, f"A{rr + 1}", "注：为计取规费等的使用，可在表中增设其中：“定额人工费”。", align=L, border=False)
        put(ws, f"H{rr + 2}", "表-08", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFGHI", [7, 15, 24, 26, 9, 11, 12, 13, 9]):
            ws.column_dimensions[col].width = w

        # ================= 表-09 综合单价分析表 =================
        ws = wb.create_sheet("表-09 综合单价分析表")
        r = 1
        for boq, result in rows:
            qty = float(boq.quantity or 1) or 1
            lab, mat, mac = float(result.labor_cost or 0), float(result.material_cost or 0), float(result.machine_cost or 0)
            mng, pro, tot = float(result.management_fee or 0), float(result.profit or 0), float(result.total or 0)
            put(ws, f"A{r}", "综合单价分析表", border=False).font = Font(bold=True, size=13)
            ws.merge_cells(f"A{r}:P{r}"); ws[f"A{r}"].alignment = C
            put(ws, f"A{r + 1}", f"工程名称：{project_name}", align=L, border=False)
            put(ws, f"G{r + 1}", f"标段：{bid_section}", align=L, border=False)
            put(ws, f"L{r + 1}", "第   页  共    页", align=C, border=False)
            put(ws, f"A{r + 2}", "项目编码", font=HDR, align=C)
            ws.merge_cells(f"A{r + 2}:B{r + 2}")
            put(ws, f"C{r + 2}", boq.code or "", align=L)
            ws.merge_cells(f"C{r + 2}:D{r + 2}")
            put(ws, f"E{r + 2}", "项目名称", font=HDR, align=C)
            ws.merge_cells(f"E{r + 2}:G{r + 2}")
            put(ws, f"H{r + 2}", boq.name or "", align=L)
            ws.merge_cells(f"H{r + 2}:J{r + 2}")
            put(ws, f"K{r + 2}", "计量单位", font=HDR, align=C)
            put(ws, f"M{r + 2}", boq.unit or "", align=C)
            put(ws, f"N{r + 2}", "工程量", font=HDR, align=C)
            put(ws, f"O{r + 2}", round(qty, 2), align=R)
            ws.merge_cells(f"O{r + 2}:P{r + 2}")
            put(ws, f"A{r + 3}", "清单综合单价组成明细", font=HDR, align=L, border=False)
            h = r + 4
            put(ws, f"A{h}", "定额\n编号", font=HDR, align=C)
            put(ws, f"B{h}", "定额项目\n名称", font=HDR, align=C)
            put(ws, f"C{h}", "定额\n单位", font=HDR, align=C)
            put(ws, f"D{h}", "数量", font=HDR, align=C)
            ws.merge_cells(f"E{h}:J{h}"); put(ws, f"E{h}", "单价", font=HDR, align=C)
            ws.merge_cells(f"K{h}:P{h}"); put(ws, f"K{h}", "合价", font=HDR, align=C)
            h2 = h + 1
            for col, v in [("E", "人工费"), ("F", "材料费"), ("H", "机械费"), ("I", "管理费\n和利润"), ("J", "风险费"),
                           ("K", "人工费"), ("M", "材料费"), ("N", "机械费"), ("O", "管理费\n和利润"), ("P", "风险费")]:
                put(ws, f"{col}{h2}", v, font=HDR, align=C)
            d = h2 + 1
            up = (lambda v: round(v / qty, 2)) if qty else (lambda v: 0)
            put(ws, f"A{d}", "1-01", align=C)
            put(ws, f"B{d}", "综合定额（组价结果）", align=L)
            put(ws, f"C{d}", boq.unit or "", align=C)
            put(ws, f"D{d}", round(qty, 2), align=R)
            for col, v in [("E", up(lab)), ("F", up(mat)), ("H", up(mac)), ("I", up(mng + pro)), ("J", 0),
                           ("K", round(lab, 2)), ("M", round(mat, 2)), ("N", round(mac, 2)), ("O", round(mng + pro, 2)), ("P", 0)]:
                put(ws, f"{col}{d}", v, align=R)
            d += 1
            put(ws, f"A{d}", "人工单价", align=L)
            put(ws, f"C{d}", "小计", align=C)
            d += 1
            put(ws, f"A{d}", "　一类人工 114 元/工日", align=L, border=False)
            put(ws, f"C{d}", "未计价材料费", align=L)
            d += 1
            put(ws, f"A{d}", "清单项目综合单价", font=HDR, align=L)
            ws.merge_cells(f"A{d}:J{d}")
            put(ws, f"K{d}", round(tot / qty, 2) if qty else 0, font=HDR, align=R)
            d += 1
            put(ws, f"A{d}", "材\n料\n费\n明\n细", font=HDR, align=C)
            for col, v in [("B", "主要材料名称、规格、型号"), ("H", "单位"), ("I", "数量"),
                           ("K", "单价（元）"), ("M", "合价（元）"), ("N", "暂估单价\n（元）"), ("O", "暂估合价\n（元）")]:
                put(ws, f"{col}{d}", v, font=HDR, align=C)
            m0 = d + 1
            put(ws, f"A{m0}", 1, align=C)
            put(ws, f"B{m0}", "主要材料（综合）", align=L)
            put(ws, f"H{m0}", boq.unit or "", align=C)
            put(ws, f"I{m0}", round(qty, 2), align=R)
            put(ws, f"K{m0}", up(mat), align=R)
            put(ws, f"M{m0}", round(mat, 2), align=R)
            put(ws, f"N{m0}", 0, align=R); put(ws, f"O{m0}", 0, align=R)
            d = m0 + 1
            put(ws, f"K{d}", "合价", font=HDR, align=C)
            put(ws, f"M{d}", round(mat, 2), align=R)
            d += 2
            put(ws, f"A{d}", "注：1.如不使用省级或行业建设主管部门发布的计价依据，可不填定额编码、名称等；\n    2.招标文件提供了暂估单价的材料，按暂估的单价填入相应栏。", align=L, border=False)
            put(ws, f"L{d}", "表-09", border=False).font = Font(size=9)
            r = d + 2
        for col, w in zip("ABCDEFGHIJKLMNOP", [9, 20, 8, 8, 9, 9, 4, 9, 9, 8, 9, 4, 9, 9, 9, 8]):
            ws.column_dimensions[col].width = w

        # ================= 表-11 总价措施项目清单与计价表 =================
        ws = wb.create_sheet("表-11 总价措施项目清单与计价表")
        put(ws, "A2", "总价措施项目清单与计价表", border=False)
        ws.merge_cells("A2:K2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "D3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "I3", "第  1  页  共  1  页", align=C, border=False)
        for addr, v in [("A4", "序号"), ("B4", "项目编码"), ("C4", "项目名称"), ("E4", "计算基础"),
                        ("F4", "费率(%)"), ("G4", "金额(元)"), ("H4", "调整费率(%)"), ("J4", "调整后金额(元)"), ("K4", "备注")]:
            put(ws, addr, v, font=HDR, align=C)
        ws.merge_cells("C4:D4")
        measures = [
            (9, "011707001001", "安全文明施工", "", "", "", ""),
            (10, "011707001002", "其中：环境保护费、文明施工费、安全施工费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            (11, "011707001003", "其中：临时设施费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            (12, "011707001004", "其中：智慧工地基础配置费", "", "", "", ""),
            (13, "011707002001", "夜间施工增加费", "20.7*10", "", "", "按实际发生计取，不发生不计取。"),
            (14, "011707004001", "二次搬运费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            (15, "011707005001", "冬雨季施工增加费", "", "", "", "不发生不计取，在编制施工招标文件时依据工程实际情况，结合市场编制费用清单。"),
            (16, "011707007001", "已完工程及设备保护费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            (17, "X011707008001", "工程定位复测、点交清理费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            (18, "X011707010001", "检验试验费", "", "", "", ""),
            (19, "X011707010002", "其中：自检试验费", "", "", "", ""),
            (20, "X011707010003", "其中：检验试验配合费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            (21, "X011707011001", "特殊地区增加费", "", "", "", "结合工程实际情况，按相关规定计取。"),
            (22, "01B991", "竣工档案编制费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", 0, ""),
        ]
        for i, (no, code, name, base, rate, amt, note) in enumerate(measures, 5):
            put(ws, f"A{i}", no, align=C)
            put(ws, f"B{i}", code, align=L)
            ws.merge_cells(f"C{i}:D{i}"); put(ws, f"C{i}", name, align=L)
            put(ws, f"E{i}", base, align=L)
            put(ws, f"F{i}", rate, align=C)
            put(ws, f"G{i}", amt, align=R)
            put(ws, f"H{i}", "", align=C); ws.merge_cells(f"H{i}:I{i}")
            put(ws, f"J{i}", "", align=R)
            put(ws, f"K{i}", note, align=L)
        last = 5 + len(measures)
        put(ws, f"A{last}", "编制人（造价人员）：", align=L, border=False)
        put(ws, f"I{last}", "复核人（造价工程师）：", align=L, border=False)
        put(ws, f"A{last + 1}", "注：1.“计算基础”中安全文明施工费可为“定额基价”、“定额人工费”或“定额人工费+定额机械费”，其他项目可为“定额人工费”或“定额人工费+定额机械费”。", align=L, border=False)
        put(ws, f"I{last + 2}", "表-11", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFGHIJK", [7, 14, 16, 16, 30, 8, 12, 8, 8, 12, 22]):
            ws.column_dimensions[col].width = w

        # ================= 表-12 其他项目清单与计价汇总表 =================
        ws = wb.create_sheet("表-12 其他项目清单与计价汇总表")
        put(ws, "A2", "其他项目清单与计价汇总表", border=False)
        ws.merge_cells("A2:G2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "C3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "F3", "第  1  页  共  1  页", align=C, border=False)
        for addr, v in [("A4", "序号"), ("B4", "项目名称"), ("D4", "金额(元)"), ("E4", "结算金额(元)"), ("G4", "备注")]:
            put(ws, addr, v, font=HDR, align=C)
        others = [
            ("1", "暂列金额", "", "明细详见表-12-1"),
            ("2", "暂估价", "", ""),
            ("2.1", "材料暂估价", "", "明细详见表-12-2"),
            ("2.2", "专业工程暂估价", "", "明细详见表-12-3"),
            ("2.3", "施工技术专项措施项目暂估价", "", "明细详见表-12-7"),
            ("3", "计日工", "", "明细详见表-12-4"),
            ("4", "总承包服务费", "", "明细详见表-12-5"),
        ]
        for i, (no, name, amt, note) in enumerate(others, 5):
            put(ws, f"A{i}", no, align=C)
            ws.merge_cells(f"B{i}:C{i}"); put(ws, f"B{i}", name, align=L)
            put(ws, f"D{i}", amt, align=R)
            put(ws, f"E{i}", "", align=R); ws.merge_cells(f"E{i}:F{i}")
            put(ws, f"G{i}", note, align=L)
        put(ws, "A12", "合    计", font=HDR, align=C)
        put(ws, "D12", "0.00", align=R)
        put(ws, "G12", "—", align=C)
        put(ws, "A13", "注：材料（工程设备）暂估单价进入清单项目综合单价，此处不汇总。", align=L, border=False)
        put(ws, "F14", "表-12", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFG", [8, 18, 18, 14, 12, 12, 24]):
            ws.column_dimensions[col].width = w

        # ================= 表-12-1 暂列金额明细表 =================
        ws = wb.create_sheet("表-12-1 暂列金额明细表")
        put(ws, "A2", "暂列金额明细表", border=False)
        ws.merge_cells("A2:E2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "C3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "F3", "第 1 页  共 1 页", align=C, border=False)
        for addr, v in [("A4", "序号"), ("B4", "项目名称"), ("D4", "计量单位"), ("E4", "暂定金额(元)"), ("G4", "备注")]:
            put(ws, addr, v, font=HDR, align=C)
        zl = [(1, "工程造价上涨", "项"), (2, "建筑工程", "项"), (3, "装饰工程", "项")]
        for i, (no, name, unit) in enumerate(zl, 5):
            put(ws, f"A{i}", no, align=C)
            ws.merge_cells(f"B{i}:C{i}"); put(ws, f"B{i}", name, align=L)
            put(ws, f"D{i}", unit, align=C)
            put(ws, f"E{i}", "", align=R); ws.merge_cells(f"E{i}:F{i}")
            put(ws, f"G{i}", "", align=L)
        put(ws, "A8", "合    计", font=HDR, align=C)
        ws.merge_cells("A8:D8")
        put(ws, "E8", "0.00", align=R); ws.merge_cells("E8:F8")
        put(ws, "G8", "—", align=C)
        put(ws, "A9", "注：此表由招标人填写，如不能详列，也可只列暂列金额总额，投标人应将上述暂列金额计入投标总价中。", align=L, border=False)
        put(ws, "F10", "表—12—1", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFG", [8, 18, 18, 12, 14, 12, 22]):
            ws.column_dimensions[col].width = w

        # ================= 表-12-3 专业工程暂估价及结算价表 =================
        ws = wb.create_sheet("表-12-3 专业工程暂估价及结算价表")
        put(ws, "A2", "专业工程暂估价及结算价表", border=False)
        ws.merge_cells("A2:I2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "D3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "H3", "第  1  页  共  1  页", align=C, border=False)
        for addr, v in [("A4", "序号"), ("B4", "工程名称"), ("C4", "工程内容"), ("E4", "暂估金额（元）"),
                        ("F4", "结算金额(元)"), ("G4", "差额±(元)"), ("I4", "备注")]:
            put(ws, addr, v, font=HDR, align=C)
        put(ws, "A5", "1", align=C)
        put(ws, "B5", "地基处理", align=L)
        ws.merge_cells("C5:D5"); put(ws, "C5", "CFG桩施工", align=L)
        put(ws, "E5", "", align=R); put(ws, "F5", "", align=R)
        ws.merge_cells("G5:H5"); put(ws, "G5", "", align=R)
        put(ws, "I5", "", align=L)
        put(ws, "A8", "合    计", font=HDR, align=C); ws.merge_cells("A8:D8")
        put(ws, "E8", "0.00", align=R); put(ws, "F8", "", align=R)
        ws.merge_cells("G8:H8"); put(ws, "G8", "", align=R)
        put(ws, "I8", "—", align=C)
        put(ws, "A9", "注：此表“暂估金额”由招标人填写，投标人应将“暂估金额”计入投标总价中。结算时按合同约定结算金额填写。", align=L, border=False)
        put(ws, "H10", "表—12—3", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFGHI", [8, 16, 16, 14, 14, 14, 12, 12, 20]):
            ws.column_dimensions[col].width = w

        # ================= 表-12-4 计日工表 =================
        ws = wb.create_sheet("表-12-4 计日工表")
        put(ws, "A2", "计 日 工 表", border=False)
        ws.merge_cells("A2:J2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "C3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "I3", "第  1  页  共  1  页", align=C, border=False)
        for addr, v in [("A4", "编号"), ("B4", "项目名称"), ("D4", "单位"), ("E4", "暂定数量"),
                        ("F4", "实际数量"), ("G4", "综合单价(元)"), ("H4", "合价")]:
            put(ws, addr, v, font=HDR, align=C)
        put(ws, "H5", "暂定", font=HDR, align=C)
        put(ws, "J5", "实际", font=HDR, align=C)
        jr = [
            ("1", "人工", "", "", "", "", ""),
            ("1.1", "土建综合工日", "工日", "", "", "", ""),
            ("人工小计", "", "", "", "", "", ""),
            ("2", "材料", "", "", "", "", ""),
            ("2.1", "钢筋 HPB300 Φ10", "t", "", "", "", ""),
            ("2.2", "水泥42.5", "t", "", "", "", ""),
            ("材料小计", "", "", "", "", "", ""),
            ("3", "机械", "", "", "", "", ""),
            ("3.1", "自升式塔吊起重机", "台班", "", "", "", ""),
            ("机械小计", "", "", "", "", "", ""),
        ]
        for i, (no, name, unit, q1, q2, price, hj) in enumerate(jr, 6):
            put(ws, f"A{i}", no, align=C)
            ws.merge_cells(f"B{i}:C{i}"); put(ws, f"B{i}", name, align=L)
            put(ws, f"D{i}", unit, align=C)
            put(ws, f"E{i}", q1, align=R)
            put(ws, f"F{i}", q2, align=R)
            put(ws, f"G{i}", price, align=R)
            put(ws, f"H{i}", hj, align=R); ws.merge_cells(f"H{i}:I{i}")
            put(ws, f"J{i}", "", align=R)
        put(ws, "A17", "4.企业管理费和利润", align=L)
        put(ws, "A18", "总    计", font=HDR, align=C)
        put(ws, "H18", "0.00", align=R); ws.merge_cells("H18:I18")
        put(ws, "A19", "注：此表项目名称、暂定数量由招标人填写，编制招标控制价时，单价由招标人按有关计价规定确定；投标时，单价由投标人自主报价。", align=L, border=False)
        put(ws, "I20", "表—12—4", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFGHIJ", [8, 16, 12, 10, 11, 11, 12, 12, 8, 10]):
            ws.column_dimensions[col].width = w

        # ================= 表-12-5 总承包服务费计价表 =================
        ws = wb.create_sheet("表-12-5 总承包服务费计价表")
        put(ws, "A2", "总承包服务费计价表", border=False)
        ws.merge_cells("A2:I2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "C3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "H3", "第  1  页  共  1  页", align=C, border=False)
        for addr, v in [("A4", "序号"), ("B4", "项目名称"), ("D4", "项目价值(元)"), ("E4", "服务内容"),
                        ("F4", "计算基础"), ("G4", "费率(%)"), ("I4", "金额(元)")]:
            put(ws, addr, v, font=HDR, align=C)
        zcb = [
            ("1", "专业发包工程管理费", "", "施工质量、进度管理；竣工资料管理", "", "", ""),
            ("2", "甲供材料设备保管费", "", "", "", "", ""),
        ]
        for i, (no, name, val, svc, base, rate, amt) in enumerate(zcb, 5):
            put(ws, f"A{i}", no, align=C)
            ws.merge_cells(f"B{i}:C{i}"); put(ws, f"B{i}", name, align=L)
            put(ws, f"D{i}", val, align=R)
            put(ws, f"E{i}", svc, align=L)
            put(ws, f"F{i}", base, align=L)
            put(ws, f"G{i}", rate, align=C)
            put(ws, f"I{i}", amt, align=R)
        put(ws, "A8", "合    计", font=HDR, align=C); ws.merge_cells("A8:B8")
        put(ws, "I8", "0.00", align=R)
        put(ws, "A9", "注：此表项目名称、服务内容由招标人填写，编制招标控制价时，费率及金额由招标人按有关计价规定确定；投标时，费率及金额由投标人自主报价。", align=L, border=False)
        put(ws, "H10", "表—12—5", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFGHI", [8, 16, 12, 14, 26, 20, 8, 10, 12]):
            ws.column_dimensions[col].width = w

        # ================= 表-13 规费、税金项目清单与计价表 =================
        ws = wb.create_sheet("表-13 规费税金项目清单与计价表")
        put(ws, "A2", "规费、税金项目计价表", border=False)
        ws.merge_cells("A2:H2"); ws["A2"].font = Font(bold=True, size=14); ws["A2"].alignment = C
        put(ws, "A3", f"工程名称：{project_name}", align=L, border=False)
        put(ws, "C3", f"标段：{bid_section}", align=L, border=False)
        put(ws, "F3", "第  1  页  共  1  页", align=C, border=False)
        for addr, v in [("A4", "序号"), ("B4", "项目名称"), ("D4", "计算基础"), ("E4", "计算基数"), ("G4", "计算费率(%)"), ("H4", "金额（元）")]:
            put(ws, addr, v, font=HDR, align=C)
        gf = [
            ("1", "规费", "社会保险费+住房公积金", "", "", ""),
            ("1.1", "社会保险费", "养老保险费+失业保险费+医疗保险费+工伤保险费", "", "", ""),
            ("1.11", "养老保险费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            ("1.12", "失业保险费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            ("1.13", "医疗保险费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            ("1.14", "工伤保险费", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            ("1.2", "住房公积金", "分部分项人工费+分部分项机械费+单价措施项目人工费+单价措施项目机械费", "", "", ""),
            ("2", "税金", "税前工程造价", "", "", round(tax, 2)),
        ]
        for i, (no, name, base, js, rate, amt) in enumerate(gf, 5):
            put(ws, f"A{i}", no, align=C)
            ws.merge_cells(f"B{i}:C{i}"); put(ws, f"B{i}", name, align=L)
            put(ws, f"D{i}", base, align=L)
            put(ws, f"E{i}", js, align=R); ws.merge_cells(f"E{i}:F{i}")
            put(ws, f"G{i}", rate, align=C)
            put(ws, f"H{i}", amt, align=R)
        put(ws, "A13", "合计", font=HDR, align=C); ws.merge_cells("A13:G13")
        put(ws, "H13", round(regulatory + tax, 2), font=HDR, align=R)
        put(ws, "A14", "编制人（造价人员）：", align=L, border=False)
        put(ws, "F14", "复核人（造价工程师）：", align=L, border=False)
        put(ws, "F15", "表-13", border=False).font = Font(size=9)
        for col, w in zip("ABCDEFGH", [8, 14, 14, 30, 12, 12, 11, 13]):
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()

def export_diff_report(snap_a, snap_b) -> bytes:
    """Generate an Excel diff report comparing two snapshots."""
    from app.models.snapshot import Snapshot  # avoid circular import

    report = diff_snapshots(snap_a, snap_b)

    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        ws.title = "差异对比表"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "版本差异对比表"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center

        ws.merge_cells("A2:G2")
        ws["A2"] = (
            f"快照 A (ID {report.snapshot_a_id}) vs 快照 B (ID {report.snapshot_b_id})  |  "
            f"生成时间: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        )
        ws["A2"].font = Font(size=9, italic=True)

        # Summary
        ws.cell(row=3, column=1, value="旧合计").font = header_font
        ws.cell(row=3, column=2, value=report.old_grand_total)
        ws.cell(row=3, column=3, value="新合计").font = header_font
        ws.cell(row=3, column=4, value=report.new_grand_total)
        ws.cell(row=3, column=5, value="差额").font = header_font
        ws.cell(row=3, column=6, value=report.grand_total_delta)

        # Headers
        headers = ["编码", "名称", "变更类型", "旧金额", "新金额", "差额"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        for row_idx, line in enumerate(report.lines, 6):
            ws.cell(row=row_idx, column=1, value=line.boq_code)
            ws.cell(row=row_idx, column=2, value=line.boq_name)
            ws.cell(row=row_idx, column=3, value=line.change_type)
            ws.cell(row=row_idx, column=4, value=line.old_total)
            ws.cell(row=row_idx, column=5, value=line.new_total)
            ws.cell(row=row_idx, column=6, value=line.delta)

        col_widths = [12, 20, 10, 14, 14, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    finally:
        wb.close()
