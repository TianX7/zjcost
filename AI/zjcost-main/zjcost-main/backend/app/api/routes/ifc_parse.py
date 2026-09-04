"""IFC model parsing API: upload IFC file, poll for parsed results, export to Excel, save to project."""

from __future__ import annotations

import io
import logging
import multiprocessing
import os
import queue
import re
import sys
import threading
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.db.session import session_scope
from app.models.boq_item import BoqItem
from app.models.line_item_quota_binding import LineItemQuotaBinding
from app.models.quota_item import QuotaItem
from app.schemas.ifc_parse import (
    IfcBoqSuggestionOut,
    IfcElementOut,
    IfcTaskStatusResponse,
    SaveToProjectRequest,
    SaveToProjectResponse,
)
from app.services.drawing_valuation_service import (
    binding_coefficient_for_units,
    create_valuation_from_drawing,
    division_for_boq,
    match_quota_for_boq,
)
from app.services.project_calc_service import run_project_calculation
from app.services.task_store import (
    delete_expired_background_tasks,
    load_background_task,
    save_background_task,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/ifc-parse", tags=["ifc-parse"])

_tasks: dict[str, dict] = {}
_tasks_lock = threading.Lock()
_TASK_TTL_SECONDS = 1800  # 30 minutes
_MAX_AUTO_VALUATION_SUGGESTIONS = 800
_MAX_3D_PREVIEW_ELEMENTS = 20000
_UPLOAD_READ_CHUNK_SIZE = 1024 * 1024
_ifc_parse_slots = threading.BoundedSemaphore(value=1)
_ifc_valuation_slots = threading.BoundedSemaphore(value=1)
_TASK_TYPE = "ifc_parse"


def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default


_IFC_PARSE_TIMEOUT_SECONDS = max(1, _env_int("IFC_PARSE_TIMEOUT_SECONDS", 180))
_IFC_MAX_UPLOAD_MB = max(1, _env_int("IFC_MAX_UPLOAD_MB", 150))


def _read_ifc_upload_limited(file: UploadFile) -> tuple[bytes, bool]:
    max_bytes = _IFC_MAX_UPLOAD_MB * 1024 * 1024
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = file.file.read(_UPLOAD_READ_CHUNK_SIZE)
        if not chunk:
            return b"".join(chunks), False
        total += len(chunk)
        if total > max_bytes:
            return b"", True
        chunks.append(chunk)


def _ifc_parse_mode() -> str:
    if getattr(sys, "frozen", False) or os.getenv("ZJCOST_PORTABLE", "").strip() == "1":
        return "thread"
    raw = os.getenv("ZJCOST_IFC_PARSE_MODE", "").strip().lower()
    if raw in {"process", "thread"}:
        return raw
    return "process"


def _ifc_parse_worker(file_bytes: bytes, filename: str, result_queue) -> None:
    """Run IFC parsing in an isolated process so crashes/hangs cannot take down the API."""
    try:
        from app.services.ifc_parse_service import parse_ifc_bytes

        def _progress(msg: str) -> None:
            try:
                result_queue.put(("progress", msg), block=False)
            except Exception:
                pass

        result = parse_ifc_bytes(file_bytes, filename, progress_callback=_progress)
        result_queue.put(("result", result))
    except BaseException as exc:
        try:
            result_queue.put(("error", f"IFC解析子进程异常: {exc}"))
        except Exception:
            pass


def _cleanup_expired_tasks() -> None:
    """Remove tasks older than TTL to prevent unbounded memory growth."""
    now = datetime.now(timezone.utc)
    with _tasks_lock:
        expired = [
            tid for tid, t in _tasks.items()
            if t.get("_created_at") and (now - t["_created_at"]).total_seconds() > _TASK_TTL_SECONDS
        ]
        for tid in expired:
            del _tasks[tid]
    delete_expired_background_tasks(_TASK_TYPE, _TASK_TTL_SECONDS)


def _dump_model(model: BaseModel) -> dict:
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


def _store_result(task_id: str, payload: dict) -> None:
    with _tasks_lock:
        if task_id not in _tasks:
            return
        payload["updated_at"] = datetime.now(timezone.utc).isoformat()
        _tasks[task_id].update(payload)
        snapshot = dict(_tasks[task_id])
    save_background_task(task_id, _TASK_TYPE, snapshot)


def _get_task(task_id: str) -> dict | None:
    with _tasks_lock:
        task = _tasks.get(task_id)
        if task is not None:
            return dict(task)

    task = load_background_task(task_id, _TASK_TYPE)
    if task is None:
        return None
    with _tasks_lock:
        _tasks[task_id] = dict(task)
    return dict(task)


def _valuation_progress_percent(message: str) -> int:
    if "完成" in message:
        return 100
    if "计算" in message:
        return 90
    if "整理" in message:
        return 95
    match = re.search(r"(\d+)\s*/\s*(\d+)", message)
    if match:
        current, total = int(match.group(1)), max(int(match.group(2)), 1)
        if "写入" in message or "清单" in message:
            return min(35, 8 + round(current / total * 27))
        if "匹配" in message or "定额" in message:
            return min(88, 40 + round(current / total * 48))
    if "读取" in message or "定额" in message:
        return 38
    if "创建" in message or "项目" in message:
        return 8
    return 5


def _serialize_ifc_element(element: dict) -> dict:
    return _dump_model(IfcElementOut(**{
        "id": element["id"], "type": element["type"], "label": element["label"],
        "name": element["name"], "element_type": element["element_type"],
        "predefined_type": element.get("predefined_type", ""),
        "object_type": element.get("object_type", ""),
        "description": element.get("description", ""),
        "count": element.get("count", 1),
        "material": element.get("material", ""),
        "length": element.get("length", 0), "width": element.get("width", 0),
        "height": element.get("height", 0), "thickness": element.get("thickness", 0),
        "area": element.get("area", 0), "volume": element.get("volume", 0),
        "unit": element.get("unit", ""),
        "quantity_estimate": element.get("quantity_estimate", 0),
        "confidence": element.get("confidence", 95.0),
        "pset_keys": element.get("pset_keys", []),
        "pos_x": element.get("pos_x", 0), "pos_y": element.get("pos_y", 0), "pos_z": element.get("pos_z", 0),
        "mesh_vertices": element.get("mesh_vertices", []),
        "mesh_indices": element.get("mesh_indices", []),
        "mesh_kind": element.get("mesh_kind", ""),
    }))


def _auto_valuate_ifc(task_id: str, suggestions_raw: list[dict]) -> dict:
    """Create project + BOQ items + auto-match quotas + calculate, same as drawing flow."""
    try:
        with session_scope() as db:
            def _progress(message: str) -> None:
                _store_result(task_id, {
                    "valuation_progress": message,
                    "valuation_progress_percent": _valuation_progress_percent(message),
                    "valuation_status": "processing",
                })

            suggestions = []
            for s in suggestions_raw:
                suggestions.append({
                    "suggested_code": s.get("suggested_code", ""),
                    "suggested_name": s.get("suggested_name", ""),
                    "suggested_unit": s.get("suggested_unit", ""),
                    "suggested_quantity": s.get("suggested_quantity", 0),
                    "characteristics": s.get("characteristics", ""),
                    "source_component_id": s.get("source_element_id", ""),
                })
            return create_valuation_from_drawing(
                db=db,
                boq_suggestions=suggestions,
                task_id=task_id,
                source_type="ifc",
                progress_callback=_progress,
            )
    except Exception as exc:
        logger.exception("IFC auto valuation failed task_id=%s: %s", task_id, exc)
        return {
            "project_id": None,
            "project_name": "",
            "boq_items_created": 0,
            "matched": 0,
            "skipped": 0,
            "grand_total": 0.0,
            "total_direct": 0.0,
            "items": [],
            "calc_summary": None,
            "review_items": [{
                "severity": "error",
                "category": "valuation_error",
                "message": f"自动计价失败: {exc}",
                "suggestion": "请先检查 IFC 构件清单是否完整，再在项目详情中手动补套定额。",
            }],
            "review_summary": {"errors": 1, "warnings": 0, "total": 1},
            "error": f"自动计价失败: {exc}",
        }


def _run_ifc_valuation(task_id: str) -> None:
    acquired = _ifc_valuation_slots.acquire(blocking=False)
    if not acquired:
        _store_result(task_id, {
            "valuation_status": "error",
            "valuation_error": "已有自动计价任务正在运行，请稍后再试。",
            "valuation_progress": "等待计价资源失败",
            "valuation_progress_percent": 0,
        })
        return

    try:
        with _tasks_lock:
            task = _tasks.get(task_id)
            suggestions_raw = list(task.get("boq_suggestions", [])) if task else []

        if not suggestions_raw:
            _store_result(task_id, {
                "valuation_status": "skipped",
                "valuation_error": "没有可计价的清单建议。",
                "valuation_progress": "未生成可计价清单",
                "valuation_progress_percent": 0,
            })
            return

        if len(suggestions_raw) > _MAX_AUTO_VALUATION_SUGGESTIONS:
            error = (
                f"清单建议 {len(suggestions_raw)} 条，超过自动计价上限 "
                f"{_MAX_AUTO_VALUATION_SUGGESTIONS} 条。请先导出筛选，或保存到项目后分批计价。"
            )
            _store_result(task_id, {
                "valuation_status": "error",
                "valuation_error": error,
                "valuation_progress": "自动计价已跳过",
                "valuation_progress_percent": 0,
            })
            return

        _store_result(task_id, {
            "valuation_status": "processing",
            "valuation_error": None,
            "valuation_progress": "正在创建计价项目并匹配定额...",
            "valuation_progress_percent": 5,
        })
        valuation = _auto_valuate_ifc(task_id, suggestions_raw)
        valuation_error = valuation.get("error")
        has_project = bool(valuation.get("project_id"))
        _store_result(task_id, {
            "valuation": valuation,
            "valuation_status": "done" if has_project else "error",
            "valuation_error": valuation_error,
            "valuation_progress": "自动计价完成" if has_project else (valuation_error or "自动计价未生成项目"),
            "valuation_progress_percent": 100 if has_project else 0,
        })
    except Exception as exc:
        logger.exception("IFC valuation task failed task_id=%s: %s", task_id, exc)
        _store_result(task_id, {
            "valuation_status": "error",
            "valuation_error": f"自动计价失败: {exc}",
            "valuation_progress": "自动计价失败",
            "valuation_progress_percent": 0,
        })
    finally:
        _ifc_valuation_slots.release()


def _run_ifc_parse_process(task_id: str, file_bytes: bytes, filename: str) -> None:
    acquired = _ifc_parse_slots.acquire(blocking=False)
    if not acquired:
        _store_result(task_id, {
            "status": "error",
            "error": "已有 IFC 解析任务正在运行，请等待当前任务结束后再上传。",
            "progress": "等待解析资源失败",
        })
        return

    try:
        def _progress(msg: str) -> None:
            _store_result(task_id, {"progress": msg})

        _progress("正在读取 IFC 文件结构...")
        ctx = multiprocessing.get_context("spawn")
        result_queue = ctx.Queue()
        process = ctx.Process(
            target=_ifc_parse_worker,
            args=(file_bytes, filename, result_queue),
            daemon=False,
        )
        process.start()

        result = None
        deadline = datetime.now(timezone.utc).timestamp() + _IFC_PARSE_TIMEOUT_SECONDS
        while process.is_alive():
            try:
                kind, payload = result_queue.get(timeout=0.5)
                if kind == "progress":
                    _progress(str(payload))
                elif kind == "result":
                    result = payload
                    break
                elif kind == "error":
                    raise RuntimeError(str(payload))
            except queue.Empty:
                pass

            if datetime.now(timezone.utc).timestamp() > deadline:
                process.terminate()
                process.join(timeout=5)
                _store_result(task_id, {
                    "status": "error",
                    "error": f"IFC 解析超过 {_IFC_PARSE_TIMEOUT_SECONDS} 秒，已自动终止。建议拆分模型或导出精简 IFC。",
                    "progress": "解析超时，已终止",
                })
                return

        while result is None:
            try:
                kind, payload = result_queue.get_nowait()
                if kind == "progress":
                    _progress(str(payload))
                elif kind == "result":
                    result = payload
                    break
                elif kind == "error":
                    raise RuntimeError(str(payload))
            except queue.Empty:
                break

        process.join(timeout=5)
        if result is None:
            exitcode = process.exitcode
            _store_result(task_id, {
                "status": "error",
                "error": f"IFC 解析子进程异常退出（exitcode={exitcode}），主系统未受影响。",
                "progress": "解析失败，已隔离",
            })
            return

        if result.get("error"):
            _store_result(task_id, {
                "status": "error",
                "error": result["error"],
                "diagnostics": result.get("diagnostics", []),
            })
            return

        _store_ifc_parse_success(task_id, result)
    except Exception as exc:
        logger.exception("IFC parse failed task_id=%s: %s", task_id, exc)
        _store_result(task_id, {
            "status": "error",
            "error": f"IFC解析异常: {exc}",
        })
    finally:
        _ifc_parse_slots.release()


def _store_ifc_parse_success(task_id: str, result: dict) -> None:
    elements_raw = result.get("elements", [])
    preview_raw = result.get("preview_elements", elements_raw)
    preview_elements = [_serialize_ifc_element(e) for e in preview_raw[:_MAX_3D_PREVIEW_ELEMENTS]]
    elements = [_serialize_ifc_element(e) for e in elements_raw]

    suggestions_raw = result.get("boq_suggestions", [])
    suggestions = [
        _dump_model(IfcBoqSuggestionOut(**s))
        for s in suggestions_raw
    ]

    total = result.get("total_elements", 0)
    detail_count = result.get("detail_element_count", total)
    preview_count = min(result.get("preview_element_count", len(preview_raw)), _MAX_3D_PREVIEW_ELEMENTS)
    aggregated_count = result.get("aggregated_element_count", len(elements_raw))
    mesh_count = result.get("mesh_element_count", 0)
    summary = (
        f"IFC 模型解析完成，共提取 {detail_count} 个明细构件，"
        f"3D 预览 {preview_count} 个（真实网格 {mesh_count} 个），"
        f"聚合清单 {aggregated_count} 条，生成 {len(suggestions_raw)} 条计价建议。"
    )
    valuation_status = "idle" if suggestions_raw else "skipped"
    valuation_progress = (
        "解析已完成，可按需手动生成计价项目。"
        if suggestions_raw
        else "未生成可计价清单"
    )

    _store_result(task_id, {
        "status": "done",
        "summary": summary,
        "elements": elements,
        "preview_elements": preview_elements,
        "boq_suggestions": suggestions,
        "statistics": result.get("statistics", {}),
        "diagnostics": result.get("diagnostics", []),
        "ifc_schema": result.get("schema", ""),
        "total_elements": total,
        "detail_element_count": detail_count,
        "preview_element_count": preview_count,
        "aggregated_element_count": aggregated_count,
        "mesh_element_count": mesh_count,
        "valuation": None,
        "valuation_status": valuation_status,
        "valuation_progress": valuation_progress,
        "valuation_progress_percent": 0,
        "valuation_error": None,
        "progress": "解析完成",
    })


def _run_ifc_parse_thread(task_id: str, file_bytes: bytes, filename: str) -> None:
    acquired = _ifc_parse_slots.acquire(blocking=False)
    if not acquired:
        _store_result(task_id, {
            "status": "error",
            "error": "已有 IFC 解析任务正在运行，请等待当前任务结束后再上传。",
            "progress": "等待解析资源失败",
        })
        return

    try:
        from app.services.ifc_parse_service import parse_ifc_bytes

        def _progress(msg: str) -> None:
            _store_result(task_id, {"progress": msg})

        _progress("正在打包版安全模式下解析 IFC，不会启动新的控制台窗口...")
        result = parse_ifc_bytes(file_bytes, filename, progress_callback=_progress)
        if result.get("error"):
            _store_result(task_id, {
                "status": "error",
                "error": result["error"],
                "diagnostics": result.get("diagnostics", []),
                "progress": "解析失败",
            })
            return
        _store_ifc_parse_success(task_id, result)
    except Exception as exc:
        logger.exception("IFC parse failed in thread mode task_id=%s: %s", task_id, exc)
        _store_result(task_id, {
            "status": "error",
            "error": f"IFC 解析异常: {exc}",
            "progress": "解析失败",
        })
    finally:
        _ifc_parse_slots.release()


def _run_ifc_parse(task_id: str, file_bytes: bytes, filename: str) -> None:
    if _ifc_parse_mode() == "thread":
        _run_ifc_parse_thread(task_id, file_bytes, filename)
    else:
        _run_ifc_parse_process(task_id, file_bytes, filename)


@router.post("", response_model=IfcTaskStatusResponse)
def upload_ifc_file(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith((".ifc", ".ifczip")):
        raise HTTPException(400, "仅支持 .ifc / .ifczip 格式的文件")

    _cleanup_expired_tasks()
    task_id = str(uuid.uuid4())
    with _tasks_lock:
        now = datetime.now(timezone.utc).isoformat()
        _tasks[task_id] = {
            "taskId": task_id,
            "status": "processing",
            "summary": "",
            "elements": [],
            "boq_suggestions": [],
            "statistics": {},
            "diagnostics": [],
            "ifc_schema": "",
            "total_elements": 0,
            "detail_element_count": 0,
            "preview_element_count": 0,
            "aggregated_element_count": 0,
            "mesh_element_count": 0,
            "error": None,
            "valuation": None,
            "preview_elements": [],
            "valuation_status": "idle",
            "valuation_progress": "",
            "valuation_progress_percent": 0,
            "valuation_error": None,
            "progress": "正在读取 IFC 文件...",
            "created_at": now,
            "updated_at": now,
            "timeout_seconds": _IFC_PARSE_TIMEOUT_SECONDS,
            "_created_at": datetime.now(timezone.utc),
        }
        initial_task = dict(_tasks[task_id])
    save_background_task(task_id, _TASK_TYPE, initial_task)

    file_bytes, too_large = _read_ifc_upload_limited(file)
    if too_large:
        _store_result(task_id, {
            "status": "error",
            "error": f"IFC 文件超过 {_IFC_MAX_UPLOAD_MB}MB 上限，请先拆分或压缩模型。",
            "progress": "文件过大，已拒绝解析",
        })
        with _tasks_lock:
            return _tasks[task_id]

    threading.Thread(
        target=_run_ifc_parse,
        args=(task_id, file_bytes, file.filename or "model.ifc"),
        daemon=True,
    ).start()

    return _tasks[task_id]


@router.get("/{task_id}", response_model=IfcTaskStatusResponse)
def get_ifc_parse_result(task_id: str):
    task = _get_task(task_id)
    if task is None:
        raise HTTPException(404, "任务不存在或已过期")
    return task


@router.post("/{task_id}/auto-valuate", response_model=IfcTaskStatusResponse)
def start_ifc_auto_valuation(task_id: str):
    with _tasks_lock:
        task = _tasks.get(task_id)
        if task is None:
            loaded = load_background_task(task_id, _TASK_TYPE)
            if loaded is None:
                raise HTTPException(404, "任务不存在或已过期")
            _tasks[task_id] = loaded
            task = _tasks[task_id]
        if task.get("status") != "done":
            raise HTTPException(400, "IFC 解析尚未完成")
        if task.get("valuation_status") == "processing":
            return task
        if task.get("valuation") and task.get("valuation", {}).get("project_id"):
            return task
        task["valuation_status"] = "processing"
        task["valuation_error"] = None
        task["valuation_progress"] = "自动计价任务已提交..."
        task["valuation_progress_percent"] = 3
        task["updated_at"] = datetime.now(timezone.utc).isoformat()
        snapshot = dict(task)
    save_background_task(task_id, _TASK_TYPE, snapshot)

    threading.Thread(target=_run_ifc_valuation, args=(task_id,), daemon=True).start()

    with _tasks_lock:
        return _tasks[task_id]


@router.get("/{task_id}/export")
def export_ifc_parse_excel(task_id: str):
    task = _get_task(task_id)
    if task is None:
        raise HTTPException(404, "任务不存在或已过期")
    if task.get("status") != "done":
        raise HTTPException(400, "任务尚未完成")

    excel_bytes = _export_excel(task_id, task)
    filename = f"ifc_quantities_{task_id[:8]}.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{task_id}/save-to-project", response_model=SaveToProjectResponse)
def save_ifc_to_project(task_id: str, body: SaveToProjectRequest):
    task = _get_task(task_id)
    if task is None:
        raise HTTPException(404, "任务不存在或已过期")
    if task.get("status") != "done":
        raise HTTPException(400, "任务尚未完成")

    suggestions = task.get("boq_suggestions", [])
    if not suggestions:
        raise HTTPException(400, "没有可保存的清单建议")

    valid_suggestions = [s for s in suggestions if s.get("suggested_quantity", 0) > 0]
    if not valid_suggestions:
        raise HTTPException(400, "所有清单项的工程量都是 0，无法保存")

    try:
        # 阶段一：同步写入清单项（仅 INSERT / UPDATE，足够快）
        with session_scope() as db:
            saved_names: list[str] = []
            boq_item_ids: list[int] = []
            for idx, s in enumerate(valid_suggestions, start=1):
                code = s.get("suggested_code", f"IFC{idx:03d}")
                name = s.get("suggested_name", "")
                unit = s.get("suggested_unit", "")
                quantity = s.get("suggested_quantity", 0)
                characteristics = s.get("characteristics", "")
                material = s.get("material", "")
                item = (
                    db.query(BoqItem)
                    .filter(BoqItem.project_id == body.project_id, BoqItem.code == code)
                    .first()
                )
                merged_characteristics = f"{characteristics}、材质：{material}" if material else characteristics
                if item is None:
                    item = BoqItem(
                        project_id=body.project_id,
                        code=code,
                        name=name,
                        unit=unit,
                        quantity=quantity,
                        characteristics=merged_characteristics,
                        division=division_for_boq(code, name),
                        sort_order=idx * 10,
                        is_dirty=True,
                    )
                    db.add(item)
                    db.flush()
                else:
                    item.name = name
                    item.unit = unit
                    item.quantity = quantity
                    item.characteristics = merged_characteristics
                    item.division = division_for_boq(code, name)
                    item.sort_order = item.sort_order or idx * 10
                    item.is_dirty = True
                    db.query(LineItemQuotaBinding).filter(
                        LineItemQuotaBinding.boq_item_id == item.id
                    ).delete(synchronize_session=False)
                    db.flush()
                boq_item_ids.append(item.id)
                saved_names.append(name)

            db.commit()
    except Exception as exc:
        logger.exception("IFC save-to-project write failed task_id=%s", task_id)
        raise HTTPException(500, "保存失败")

    # 阶段二：异步匹配定额并计价（重计算放后台，避免请求超时）
    project_id = body.project_id
    _store_result(task_id, {
        "valuation_status": "processing",
        "valuation_progress": f"正在保存到项目并匹配定额（共 {len(saved_names)} 条清单）...",
        "valuation_progress_percent": 5,
        "valuation_error": None,
    })
    threading.Thread(
        target=_run_ifc_save_finish,
        args=(task_id, project_id, boq_item_ids, saved_names),
        daemon=True,
    ).start()

    return SaveToProjectResponse(
        project_id=project_id,
        boq_items_created=len(saved_names),
        boq_items=saved_names,
        matched=0,
        skipped=0,
        grand_total=None,
        status="processing",
        message=f"已保存 {len(saved_names)} 条清单，正在后台匹配定额并计价...",
    )


def _run_ifc_save_finish(task_id: str, project_id: int, boq_item_ids: list[int], saved_names: list[str]) -> None:
    """后台完成 IFC 保存项目后的定额匹配与计价，结果写入任务 valuation 字段。"""
    acquired = _ifc_valuation_slots.acquire(blocking=False)
    if not acquired:
        _store_result(task_id, {
            "valuation_status": "error",
            "valuation_error": "已有计价任务正在运行，请稍后在项目详情中重新计价。",
            "valuation_progress": "等待计价资源失败",
            "valuation_progress_percent": 0,
        })
        return

    def _progress(message: str) -> None:
        _store_result(task_id, {
            "valuation_progress": message,
            "valuation_progress_percent": _valuation_progress_percent(message),
            "valuation_status": "processing",
        })

    try:
        with session_scope() as db:
            _progress("正在读取定额库...")
            quotas = db.query(QuotaItem).all()
            boq_items = (
                db.query(BoqItem).filter(BoqItem.id.in_(boq_item_ids)).all()
            )
            matched = 0
            skipped = 0
            for index, item in enumerate(boq_items, start=1):
                if index == 1 or index % 20 == 0 or index == len(boq_items):
                    _progress(f"正在匹配定额 {index}/{len(boq_items)}...")
                quota, confidence = match_quota_for_boq(item, quotas)
                if quota and confidence >= 0.3:
                    db.add(LineItemQuotaBinding(
                        boq_item_id=item.id,
                        quota_item_id=quota.id,
                        coefficient=binding_coefficient_for_units(item.unit, quota.unit),
                    ))
                    matched += 1
                else:
                    skipped += 1

            db.commit()

            grand_total: float | None = None
            if matched > 0:
                try:
                    _progress("正在计算项目造价...")
                    summary, _ = run_project_calculation(project_id=project_id, db=db, incremental=True)
                    grand_total = summary.grand_total
                except Exception:
                    logger.exception("IFC save-to-project calculation failed task_id=%s", task_id)

        _store_result(task_id, {
            "valuation_status": "done",
            "valuation": {
                "project_id": project_id,
                "boq_items_created": len(saved_names),
                "matched": matched,
                "skipped": skipped,
                "grand_total": grand_total if grand_total is not None else 0.0,
                "error": None,
            },
            "valuation_progress": "自动计价完成",
            "valuation_progress_percent": 100,
            "valuation_error": None,
        })
    except Exception as exc:
        logger.exception("IFC save-to-project finish failed task_id=%s", task_id)
        _store_result(task_id, {
            "valuation_status": "error",
            "valuation_error": "保存并计价失败，请在项目详情中重新计价。",
            "valuation_progress": "计价失败",
            "valuation_progress_percent": 0,
        })
    finally:
        _ifc_valuation_slots.release()


def _export_excel(task_id: str, task: dict) -> bytes:
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)

    # --- Sheet 0: 封面（实训手册版式） ---
    cover = wb.active
    cover.title = "封-2 工程量清单封面"
    cover.merge_cells("C1:E1"); cover["C1"] = "工程"
    cover["A2"] = "工程量清单"; cover["A2"].font = Font(bold=True, size=24); cover.merge_cells("A2:G2")
    cover["A2"].alignment = Alignment(horizontal="center", vertical="center")
    cover["A3"] = f"工程名称：IFC 模型解析工程量清单（任务 {task_id[:8]}）"; cover.merge_cells("A3:G3")
    cover["A4"] = "发  包  人：____________________"; cover.merge_cells("A4:F4")
    n_elems = len(task.get("elements", []))
    cover["A5"] = f"清单内容：解析构件 {n_elems} 类，生成构件清单、清单建议与统计表"; cover.merge_cells("A5:G5")
    cover["A6"] = f"编制时间：{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}"; cover.merge_cells("A6:G6")
    cover["E7"] = "封-2"; cover["E7"].font = Font(size=9)
    _thin = Side(style="thin", color="999999")
    _bd = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    for r in range(2, 7):
        for c in range(1, 8):
            cover.cell(row=r, column=c).border = _bd
    for col, w in zip("ABCDEFG", [14, 14, 16, 16, 16, 16, 16]):
        cover.column_dimensions[col].width = w

    # --- Sheet 00: 总说明 ---
    s_total = wb.create_sheet("表-01 总说明")
    s_total.merge_cells("A1:D1")
    s_total["A1"] = "总  说  明"; s_total["A1"].font = Font(bold=True, size=16)
    s_total["A1"].alignment = Alignment(horizontal="center", vertical="center")
    s_total["A2"] = f"工程名称：IFC 模型解析工程量清单（任务 {task_id[:8]}）"
    s_total["E4"] = "表-01"; s_total["E4"].font = Font(size=9)
    notes = [task.get("summary", "")] + (task.get("diagnostics") or [])
    for i, note in enumerate(notes):
        c = s_total.cell(row=4 + i, column=1, value=("　　" + note) if note else "")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        s_total.merge_cells(start_row=4 + i, start_column=1, end_row=4 + i, end_column=5)
    s_total.column_dimensions["A"].width = 14
    for col in "BCDE":
        s_total.column_dimensions[col].width = 30

    # --- Sheet 1: 构件清单 ---
    ws = wb.create_sheet(title="IFC构件清单")
    ws.merge_cells("A1:L1")
    ws["A1"] = "IFC模型构件解析报表"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A2"] = "任务编号"
    ws["B2"] = task_id
    ws["A3"] = "生成时间"
    ws["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A4"] = "解析摘要"
    ws["B4"] = task.get("summary", "")
    ws["B4"].alignment = wrap
    ws["A5"] = "IFC Schema"
    ws["B5"] = task.get("ifc_schema", "")

    headers = [
        "序号", "构件类型", "名称", "规格型号", "材料",
        "数量", "单位", "工程量", "长(m)", "宽(m)", "高(m)", "厚(mm)",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, item in enumerate(task.get("elements", []), 8):
        ws.cell(row=row_idx, column=1, value=row_idx - 7)
        ws.cell(row=row_idx, column=2, value=item.get("label", ""))
        ws.cell(row=row_idx, column=3, value=item.get("name", ""))
        ws.cell(row=row_idx, column=4, value=item.get("element_type", ""))
        ws.cell(row=row_idx, column=5, value=item.get("material", ""))
        ws.cell(row=row_idx, column=6, value=item.get("count", 0))
        ws.cell(row=row_idx, column=7, value=item.get("unit", ""))
        ws.cell(row=row_idx, column=8, value=item.get("quantity_estimate", 0))
        ws.cell(row=row_idx, column=9, value=item.get("length", 0))
        ws.cell(row=row_idx, column=10, value=item.get("width", 0))
        ws.cell(row=row_idx, column=11, value=item.get("height", 0))
        ws.cell(row=row_idx, column=12, value=item.get("thickness", 0) * 1000)

    widths = [6, 12, 22, 28, 16, 8, 6, 12, 8, 8, 8, 8]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    # --- Sheet 2: 清单建议 ---
    ws2 = wb.create_sheet("GB50500清单建议")
    suggestion_headers = [
        "源构件", "清单编码", "清单名称", "单位", "建议工程量",
        "项目特征", "材料", "构件数", "置信度",
    ]
    for col, header in enumerate(suggestion_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row_idx, item in enumerate(task.get("boq_suggestions", []), 2):
        ws2.cell(row=row_idx, column=1, value=item.get("source_element_id", ""))
        ws2.cell(row=row_idx, column=2, value=item.get("suggested_code", ""))
        ws2.cell(row=row_idx, column=3, value=item.get("suggested_name", ""))
        ws2.cell(row=row_idx, column=4, value=item.get("suggested_unit", ""))
        ws2.cell(row=row_idx, column=5, value=item.get("suggested_quantity", 0))
        ws2.cell(row=row_idx, column=6, value=item.get("characteristics", ""))
        ws2.cell(row=row_idx, column=7, value=item.get("material", ""))
        ws2.cell(row=row_idx, column=8, value=item.get("element_count", 0))
        ws2.cell(row=row_idx, column=9, value=item.get("confidence", 0))
        ws2.cell(row=row_idx, column=6).alignment = wrap
    for idx, width in enumerate([14, 14, 28, 8, 14, 44, 12, 8, 10], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    # --- Sheet 3: 统计 ---
    ws3 = wb.create_sheet("构件统计")
    ws3.cell(row=1, column=1, value="构件类型").font = header_font
    ws3.cell(row=1, column=2, value="数量").font = header_font
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=2).fill = header_fill
    for row_idx, (label, count) in enumerate(task.get("statistics", {}).items(), 2):
        ws3.cell(row=row_idx, column=1, value=label)
        ws3.cell(row=row_idx, column=2, value=count)
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 12

    # --- Sheet 4: 诊断 ---
    ws4 = wb.create_sheet("诊断信息")
    ws4.cell(row=1, column=1, value="诊断信息").font = header_font
    for row_idx, item in enumerate(task.get("diagnostics", []), 2):
        ws4.cell(row=row_idx, column=1, value=item)
        ws4.cell(row=row_idx, column=1).alignment = wrap
    ws4.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
