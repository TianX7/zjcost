"""Drawing recognition API: upload a drawing, then poll for parsed results."""

from __future__ import annotations

import asyncio
import logging
import threading
import hashlib
import io
import json
import math
import os
import re
import subprocess
import sys
import tempfile
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from starlette.concurrency import run_in_threadpool
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from pydantic import BaseModel, Field

from app.db.session import session_scope
from app.schemas.calc_result import ProjectCalcSummary
from app.services.drawing_recognition_service import (
    components_to_boq_suggestions,
    recognize_drawing,
)
from app.services.drawing_valuation_service import create_valuation_from_drawing
from app.services.dxf_analysis_service import analyze_dxf_bytes
from app.services.dwg_conversion_service import convert_dxf_to_dwg_bytes, get_converter_status
from app.services.task_store import load_background_task, save_background_task
from app.utils.datetime import parse_datetime
from app.utils.headers import build_attachment_disposition, sanitize_filename

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/drawing-recognition", tags=["drawing-recognition"])

_tasks: dict[str, dict] = {}
_tasks_lock = threading.Lock()

# CAD 原图渲染图单独存放：解析中通过 /{task_id}/raster 按需拉取一次，
# 避免数 MB base64 混进每次轻量轮询响应；终态后并入任务快照
_RASTER_STORE: dict[str, dict] = {}
_raster_lock = threading.Lock()

# 快速看图几何数据单独存放（几 MB 线条坐标，与渲染图同样按需拉取）
_GEOMETRY_STORE: dict[str, dict] = {}
_geometry_lock = threading.Lock()

# 嵌入式 CAD 快速看图状态：同一时间仅一个实例（task_id + 坐标文件）
_EMBED_STATE: dict = {"task_id": "", "rect_file": None, "state_file": None, "proc": None}
_embed_lock = threading.Lock()
EMBED_START_TIMEOUT = 12.0

# 同一张图纸重复上传直接复用已有结果（sha256 -> task_id），秒出
_RESULT_CACHE: dict[str, str] = {}

# 上传的原始图纸落盘目录（供"用 CAD 快速看图打开"功能），后端程序目录下
_DRAWING_CACHE_DIR = Path(__file__).resolve().parents[3] / "uploads_cache"


def _find_cad_reader_exe() -> str | None:
    """探测 CAD 快速看图内核：项目内置优先（部署零依赖）→ 注册表 → 常见安装路径。"""
    # 项目内置的 CAD 快速看图（backend/tools/CADReader/，随项目分发）
    bundled = Path(__file__).resolve().parents[3] / "tools" / "CADReader" / "CADReader.exe"
    if bundled.is_file():
        return str(bundled)
    try:
        import winreg

        for root in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
            for view in (0, winreg.KEY_WOW64_64KEY, winreg.KEY_WOW64_32KEY):
                try:
                    key = winreg.OpenKey(
                        root,
                        r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
                        0,
                        winreg.KEY_READ | view,
                    )
                except OSError:
                    continue
                for i in range(winreg.QueryInfoKey(key)[0]):
                    try:
                        sub = winreg.OpenKey(key, winreg.EnumKey(key, i))
                        try:
                            name = winreg.QueryValueEx(sub, "DisplayName")[0]
                        except OSError:
                            name = ""
                        if "CAD快速看图" in name or "CADReader" in name or "看图" in name:
                            try:
                                loc = winreg.QueryValueEx(sub, "InstallLocation")[0]
                            except OSError:
                                loc = ""
                            if loc and (Path(loc) / "CADReader.exe").exists():
                                return str(Path(loc) / "CADReader.exe")
                    except OSError:
                        continue
    except Exception:
        pass
    for candidate in (
        r"D:\CADReader\CADReader.exe",
        r"C:\Program Files\Glodon\CADReader\CADReader.exe",
        r"C:\Program Files (x86)\Glodon\CADReader\CADReader.exe",
        r"C:\CADReader\CADReader.exe",
    ):
        if Path(candidate).exists():
            return candidate
    return None
_drawing_valuation_slots = threading.BoundedSemaphore(value=1)
_UPLOAD_READ_CHUNK_SIZE = 1024 * 1024
_TASK_TYPE = "drawing_recognition"
# 内存中任务缓存 TTL：7 天，避免 _tasks 字典无限增长
_TASK_TTL_SECONDS = 7 * 24 * 60 * 60


def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default


def _max_upload_bytes() -> int:
    if "DRAWING_MAX_UPLOAD_MB" in os.environ:
        return max(1, _env_int("DRAWING_MAX_UPLOAD_MB", 100)) * 1024 * 1024
    return max(1, _env_int("MAX_UPLOAD_SIZE", 100 * 1024 * 1024))


def _max_auto_valuation_suggestions() -> int:
    return max(1, _env_int("DRAWING_MAX_AUTO_VALUATION_SUGGESTIONS", 1200))


def _reject_if_too_large(file_bytes: bytes) -> None:
    max_bytes = _max_upload_bytes()
    if len(file_bytes) > max_bytes:
        limit_mb = max_bytes // (1024 * 1024)
        raise HTTPException(status_code=413, detail=f"图纸文件超过 {limit_mb}MB 上限，请压缩或拆分后再上传")


async def _read_upload_bytes_limited(file: UploadFile) -> bytes:
    max_bytes = _max_upload_bytes()
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(_UPLOAD_READ_CHUNK_SIZE)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            limit_mb = max_bytes // (1024 * 1024)
            raise HTTPException(
                status_code=413,
                detail=(
                    f"\u56fe\u7eb8\u6587\u4ef6\u8d85\u8fc7 {limit_mb}MB "
                    "\u4e0a\u9650\uff0c\u8bf7\u538b\u7f29\u6216\u62c6\u5206\u540e\u518d\u4e0a\u4f20"
                ),
            )
        chunks.append(chunk)
    return b"".join(chunks)


class ComponentOut(BaseModel):
    id: str
    type: str
    count: int
    spec: str
    confidence: float
    material: str = ""
    unit: str = ""
    quantity_estimate: float = 0.0
    length_m: float = 0.0
    area_m2: float = 0.0
    layers: list[str] = Field(default_factory=list)
    calc_note: str = ""


class BoqSuggestionOut(BaseModel):
    source_component_id: str
    suggested_code: str
    suggested_name: str
    suggested_unit: str
    suggested_quantity: float
    characteristics: str
    confidence: float
    material: str = ""
    component_count: int = 0


class LayerSummaryOut(BaseModel):
    layer: str
    count: int
    classified_as: str = ""
    entity_types: dict[str, int] = Field(default_factory=dict)


class DrawingValuationItemOut(BaseModel):
    boq_item_id: int
    code: str
    name: str
    unit: str
    quantity: float
    quota_item_id: int | None = None
    quota_code: str = ""
    quota_name: str = ""
    match_confidence: float = 0.0
    match_reason: str = ""
    match_reasons: list[str] = Field(default_factory=list)
    status: str = "skipped"
    total: float = 0.0


class DrawingValuationReviewItemOut(BaseModel):
    severity: str = "warning"
    category: str = ""
    message: str = ""
    suggestion: str = ""
    boq_item_id: int | None = None
    code: str = ""
    name: str = ""


class DrawingValuationOut(BaseModel):
    project_id: int | None = None
    project_name: str = ""
    boq_items_created: int = 0
    matched: int = 0
    skipped: int = 0
    grand_total: float = 0.0
    total_direct: float = 0.0
    items: list[DrawingValuationItemOut] = Field(default_factory=list)
    calc_summary: ProjectCalcSummary | None = None
    review_items: list[DrawingValuationReviewItemOut] = Field(default_factory=list)
    review_summary: dict[str, int] = Field(default_factory=dict)
    error: Optional[str] = None


class TaskStatusResponse(BaseModel):
    taskId: str
    status: str
    drawing_type: str = ""
    summary: str = ""
    components: list[ComponentOut] = Field(default_factory=list)
    boq_suggestions: list[BoqSuggestionOut] = Field(default_factory=list)
    diagnostics: list[str] = Field(default_factory=list)
    layer_summary: list[LayerSummaryOut] = Field(default_factory=list)
    disciplines: list[dict] = Field(default_factory=list)
    quality_score: Optional[dict] = None
    preview_svg: str = ""
    preview_svg_hd: str = ""
    cad_geometry: Optional[dict] = None
    cad_raster: Optional[dict] = None
    cad_raster_ready: bool = False
    cad_geometry_ready: bool = False
    valuation: DrawingValuationOut | None = None
    valuation_status: str = "idle"
    valuation_progress: str = ""
    valuation_progress_percent: int = 0
    valuation_error: Optional[str] = None
    progress: str = ""
    progress_percent: Optional[int] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None
    error: Optional[str] = None


class ConverterCandidateOut(BaseModel):
    name: str
    kind: str
    source: str
    bundled: bool = False


class ConverterStatusResponse(BaseModel):
    dxf_to_dwg: bool
    dwg_to_dxf: bool
    candidates: dict[str, list[ConverterCandidateOut]]
    bundled_dirs: list[str] = Field(default_factory=list)
    timeout_seconds: int = 0
    instructions: str = ""


def _dump_model(model: BaseModel) -> dict:
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()


def _json_safe(value):
    if isinstance(value, float):
        return value if math.isfinite(value) else 0.0
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    return value


def _cleanup_expired_tasks() -> None:
    """清理内存中超过 TTL 的已完成/失败任务，避免无限增长。"""
    cutoff = datetime.now(timezone.utc).timestamp() - _TASK_TTL_SECONDS
    with _tasks_lock:
        expired = [
            task_id
            for task_id, task in _tasks.items()
            if task.get("status") in ("done", "error")
            and (parse_datetime(task.get("updated_at", task.get("created_at"))) or datetime(1970, 1, 1, tzinfo=timezone.utc)).timestamp() < cutoff
        ]
        for task_id in expired:
            _tasks.pop(task_id, None)
            with _raster_lock:
                _RASTER_STORE.pop(task_id, None)
            with _geometry_lock:
                _GEOMETRY_STORE.pop(task_id, None)


def _store_result(task_id: str, payload: dict) -> None:
    _cleanup_expired_tasks()
    with _tasks_lock:
        if task_id not in _tasks:
            return
        payload["updated_at"] = datetime.now(timezone.utc).isoformat()
        _tasks[task_id].update(_json_safe(payload))
        snapshot = dict(_tasks[task_id])
    save_background_task(task_id, _TASK_TYPE, snapshot)


def _get_task(task_id: str) -> dict | None:
    _cleanup_expired_tasks()
    with _tasks_lock:
        task = _tasks.get(task_id)
        if task is not None:
            return dict(task)

    task = load_background_task(task_id, _TASK_TYPE)
    if task is None:
        return None
    with _tasks_lock:
        _tasks[task_id] = dict(task)
    return dict(task)


def _valuation_progress_percent(message: str) -> int:
    if "完成" in message:
        return 100
    if "计算项目造价" in message:
        return 90
    if "整理结果" in message:
        return 95
    match = re.search(r"写入清单项\s+(\d+)/(\d+)", message)
    if match:
        current, total = int(match.group(1)), max(int(match.group(2)), 1)
        return min(35, 8 + round(current / total * 27))
    match = re.search(r"匹配定额\s+(\d+)/(\d+)", message)
    if match:
        current, total = int(match.group(1)), max(int(match.group(2)), 1)
        return min(88, 40 + round(current / total * 48))
    if "读取定额库" in message:
        return 38
    if "创建计价项目" in message:
        return 8
    return 5


def _safe_report_filename(task_id: str, suffix: str) -> str:
    safe_id = "".join(ch for ch in task_id if ch.isalnum() or ch == "-")[:36] or "drawing"
    return f"drawing_quantities_{safe_id}.{suffix}"


def _auto_valuate_suggestions(task_id: str, suggestions: list[dict], progress_callback=None) -> dict:
    try:
        with session_scope() as db:
            return create_valuation_from_drawing(
                db=db,
                boq_suggestions=suggestions,
                task_id=task_id,
                progress_callback=progress_callback,
            )
    except Exception as exc:
        logger.exception("Drawing auto valuation failed task_id=%s: %s", task_id, exc)
        return {
            "project_id": None,
            "project_name": "",
            "boq_items_created": 0,
            "matched": 0,
            "skipped": 0,
            "grand_total": 0.0,
            "total_direct": 0.0,
            "items": [],
            "calc_summary": None,
            "error": f"自动计价失败: {exc}",
        }


def _run_drawing_valuation(task_id: str) -> None:
    acquired = _drawing_valuation_slots.acquire(blocking=False)
    if not acquired:
        _store_result(task_id, {
            "valuation_status": "error",
            "valuation_error": "已有自动计价任务正在运行，识别结果已保留。请稍后在项目中重新计价或重新上传。",
            "valuation_progress": "等待计价资源失败",
            "valuation_progress_percent": 0,
        })
        return

    try:
        with _tasks_lock:
            task = _tasks.get(task_id)
            suggestions = list(task.get("boq_suggestions", [])) if task else []

        if not suggestions:
            _store_result(task_id, {
                "valuation_status": "skipped",
                "valuation_error": "没有可计价的清单建议。",
                "valuation_progress": "未生成可计价清单",
                "valuation_progress_percent": 0,
            })
            return

        limit = _max_auto_valuation_suggestions()
        if len(suggestions) > limit:
            error = f"清单建议 {len(suggestions)} 条，超过自动计价上限 {limit} 条。请先导出筛选，或保存项目后分批计价。"
            _store_result(task_id, {
                "valuation_status": "error",
                "valuation_error": error,
                "valuation_progress": "自动计价已跳过",
                "valuation_progress_percent": 0,
            })
            return

        def _progress(message: str) -> None:
            _store_result(task_id, {
                "valuation_status": "processing",
                "valuation_progress": message,
                "valuation_progress_percent": _valuation_progress_percent(message),
            })

        _store_result(task_id, {
            "valuation_status": "processing",
            "valuation_error": None,
            "valuation_progress": "正在创建计价项目并匹配定额...",
            "valuation_progress_percent": 5,
        })
        valuation = _auto_valuate_suggestions(task_id, suggestions, _progress)
        valuation_error = valuation.get("error")
        has_project = bool(valuation.get("project_id"))
        _store_result(task_id, {
            "valuation": valuation,
            "valuation_status": "done" if has_project else "error",
            "valuation_error": valuation_error,
            "valuation_progress": "自动计价完成" if has_project else (valuation_error or "自动计价未生成项目"),
            "valuation_progress_percent": 100 if has_project else 0,
        })
    except Exception as exc:
        logger.exception("Drawing valuation task failed task_id=%s: %s", task_id, exc)
        _store_result(task_id, {
            "valuation_status": "error",
            "valuation_error": f"自动计价失败: {exc}",
            "valuation_progress": "自动计价失败",
            "valuation_progress_percent": 0,
        })
    finally:
        _drawing_valuation_slots.release()


def _start_drawing_valuation(task_id: str) -> None:
    threading.Thread(target=_run_drawing_valuation, args=(task_id,), daemon=True).start()


def _shixun_cn_upper(amount: float) -> str:
    """数字金额转中文大写（演示精度到分）。"""
    units = ["", "拾", "佰", "仟", "万", "拾万", "佰万", "仟万", "亿"]
    digits = "零壹贰叁肆伍陆柒捌玖"
    if amount is None or amount <= 0:
        return "零元整"
    try:
        amt = int(round(float(amount) * 100))
    except (TypeError, ValueError):
        return "零元整"
    yuan = amt // 100
    jiao = (amt % 100) // 10
    fen = amt % 10
    if yuan == 0:
        out = ""
    else:
        parts = []
        i = 0
        while yuan:
            n = yuan % 10
            if n:
                parts.insert(0, digits[n] + units[i])
            elif parts and (not parts[0] or parts[0][0] != "零"):
                parts.insert(0, "零")
            yuan //= 10
            i += 1
        out = "".join(parts) + "元"
    if jiao == 0 and fen == 0:
        return out + "整" if out else "零元整"
    if jiao:
        out += digits[jiao] + "角"
    if fen:
        out += digits[fen] + "分"
    return out


def _shixun_border(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if c <= 1 else "left",
                vertical="center",
                wrap_text=(c in (3, 4)),
            )


def _shixun_title(ws, row: int, end_col: int, text: str, size: int = 16, bold: bool = True) -> None:
    ws.cell(row=row, column=1, value=text).font = Font(bold=bold, size=size)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")


def _shixun_header(ws, row: int, headers: list[str], fill: str = "D9EAF7") -> None:
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _build_shixun_workbook_bytes(task_id: str, task: dict) -> bytes:
    wb = openpyxl.Workbook()
    valuation = task.get("valuation") or {}
    items = valuation.get("items", []) or []
    suggestions = task.get("boq_suggestions", []) or []
    grand = float(valuation.get("grand_total") or 0)
    subtotal = float(sum((it.get("total") or 0) for it in items)) if items else 0
    tax = round(grand * 0.09, 2) if grand else 0
    summary = (task.get("summary") or "识别完成").strip()
    project_name = f"图纸智能识别工程量报表（任务 {task_id[:8]}）"
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")

    # ================= 封-2 招标控制价封面 =================
    ws = wb.active
    ws.title = "封-2 招标控制价封面"
    ws.merge_cells("C1:E1"); ws["C1"] = "工程"
    ws["A2"] = "招标控制价"; ws["A2"].font = Font(bold=True, size=24); ws.merge_cells("A2:G2")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"] = f"工程名称：{project_name}"; ws.merge_cells("A3:G3")
    ws["A4"] = "招  标  人：____________________"; ws.merge_cells("A4:F4")
    ws["A5"] = f"招标控制价（小写）：{grand:,.2f} 元"; ws.merge_cells("A5:G5")
    ws["A6"] = f"（大写）：{_shixun_cn_upper(grand)}"; ws.merge_cells("A6:G6")
    ws["A7"] = f"编制时间：{generated}"; ws.merge_cells("A7:G7")
    ws["E8"] = "封-2"; ws["E8"].font = Font(size=9)
    _shixun_border(ws, 2, 6, 1, 7)
    for col, w in zip("ABCDEFG", [14, 14, 16, 16, 16, 16, 16]):
        ws.column_dimensions[col].width = w

    # ================= 扉-2 招标控制价扉页 =================
    ws = wb.create_sheet("扉-2 招标控制价扉页")
    ws["A1"] = "工程"; ws.merge_cells("C1:G1")
    ws["C2"] = "招标控制价"; ws.merge_cells("C2:G2"); ws["C2"].font = Font(bold=True, size=20)
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"] = "招标控制价（小写）："; ws.merge_cells("A3:C3")
    ws["D3"] = f"{grand:,.2f} 元"; ws.merge_cells("D3:G3")
    ws["A4"] = "（大写）："; ws.merge_cells("A4:C4")
    ws["D4"] = _shixun_cn_upper(grand); ws.merge_cells("D4:G4")
    ws["A5"] = f"项目名称：{project_name}"; ws.merge_cells("A5:G5")
    ws["A6"] = f"编制时间：{generated}"; ws.merge_cells("A6:G6")
    ws["H6"] = "扉—2"; ws["H6"].font = Font(size=9)
    _shixun_border(ws, 3, 6, 1, 7)
    for col, w in zip("ABCDEFGH", [14, 14, 16, 16, 16, 16, 16, 8]):
        ws.column_dimensions[col].width = w

    # ================= 表-01 总说明 =================
    ws = wb.create_sheet("表-01 总说明")
    _shixun_title(ws, 1, 4, "总  说  明", size=16)
    ws["A2"] = f"工程名称：{project_name}"; ws.merge_cells("A2:D2")
    ws["E5"] = "表-01"; ws["E5"].font = Font(size=9)
    notes = [summary] + (task.get("diagnostics") or [])
    for i, note in enumerate(notes):
        ws.cell(row=4 + i, column=1, value=("　" * 2 + note if i > 0 else f"　　{summary}")).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        ws.merge_cells(start_row=4 + i, start_column=1, end_row=4 + i, end_column=5)
    _shixun_border(ws, 4, 4 + max(len(notes), 1) - 1, 1, 5)
    ws.cell(row=4, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    for idx, w in enumerate([14, 30, 30, 30, 10], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # ================= 表-04 单位工程招标控制价汇总表 =================
    ws = wb.create_sheet("表-04 单位工程招标控制价汇总表")
    _shixun_title(ws, 2, 5, "单位工程招标控制价汇总表", size=14)
    ws["A3"] = f"工程名称：{project_name}"; ws.merge_cells("A3:C3")
    ws["E3"] = "第  1  页  共  1  页"
    _shixun_header(ws, 4, ["序号", "汇总内容", "金额（元）", "其中：暂估价（元）", "备注"])
    rows4 = [
        ("一", "分部分项工程费", subtotal, 0, ""),
        ("二", "措施项目费", 0, 0, "演示暂缺"),
        ("三", "其他项目费", 0, 0, "演示暂缺"),
        ("四", "规费", 0, 0, ""),
        ("五", "税金", tax, 0, "按 9% 演示"),
        ("", "招标控制价合计", grand, 0, ""),
    ]
    for i, (seq, name, val, temp, note) in enumerate(rows4, 5):
        ws.cell(row=i, column=1, value=seq)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=round(val, 2) if val else 0)
        ws.cell(row=i, column=4, value=temp)
        ws.cell(row=i, column=5, value=note)
    _shixun_border(ws, 4, 4 + len(rows4), 1, 5)
    for col, w in zip("ABCDEF", [8, 26, 16, 18, 30, 14]):
        ws.column_dimensions[col].width = w

    # ================= 表-08 分部分项工程和单价措施项目清单与计价表 =================
    ws = wb.create_sheet("表-08 分部分项工程清单与计价表")
    _shixun_title(ws, 2, 7, "分部分项工程和单价措施项目清单与计价表", size=14)
    ws["A3"] = f"工程名称：{project_name}"
    _shixun_header(ws, 4, ["序号", "项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "综合单价", "合价", "备注"])
    ws.merge_cells("G4:I4")
    empty_note = "识别结果不含有效清单项，金额按 0 填报" if not items and not suggestions else ""
    data_rows = []
    if items:
        for it in items:
            qty = float(it.get("quantity") or 1) or 1
            data_rows.append(
                (it.get("code") or "", it.get("name") or "", "", it.get("unit") or "",
                 float(it.get("quantity") or 0), round(float(it.get("total") or 0) / qty, 2), float(it.get("total") or 0))
            )
    elif suggestions:
        for s in suggestions:
            data_rows.append(
                (s.get("suggested_code") or "", s.get("suggested_name") or "",
                 s.get("characteristics") or f"材料：{s.get('material') or ''}", s.get("suggested_unit") or "",
                 float(s.get("suggested_quantity") or 0), 0, 0)
            )
    if not data_rows:
        data_rows.append(("", "——", "（空）", "项", 0, 0, 0))
    for i, (code, name, feat, unit, qty, price, total) in enumerate(data_rows, 5):
        ws.cell(row=i, column=1, value=i - 4)
        ws.cell(row=i, column=2, value=code)
        ws.cell(row=i, column=3, value=name)
        ws.cell(row=i, column=4, value=feat or empty_note)
        ws.cell(row=i, column=5, value=unit)
        ws.cell(row=i, column=6, value=qty)
        ws.cell(row=i, column=7, value=price)
        ws.cell(row=i, column=8, value=total)
        ws.cell(row=i, column=9, value="")
    _shixun_border(ws, 4, 4 + len(data_rows), 1, 9)
    for idx, w in enumerate([8, 16, 26, 40, 10, 12, 12, 14, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # ================= 表-09 综合单价分析表（每清单项一块） =================
    ws = wb.create_sheet("表-09 综合单价分析表")
    _shixun_title(ws, 1, 10, "综合单价分析表", size=14)
    ws["A2"] = f"工程名称：{project_name}"; ws.merge_cells("A2:F2")
    _shixun_header(ws, 5, ["定额编号", "定额项目名称", "定额单位", "数量", "单价", "合价"])
    row = 6
    if not items:
        ws.cell(row=row, column=2, value="识别结果不含可计价清单项，按 0 填报")
        _shixun_border(ws, 6, 6, 1, 10)
    else:
        for it in items:
            ws.cell(row=row, column=1, value="项目编码"); ws.cell(row=row, column=3, value=it.get("code") or "")
            ws.cell(row=row, column=5, value="项目名称"); ws.cell(row=row, column=7, value=it.get("name") or "")
            ws.cell(row=row, column=9, value="计量单位"); ws.cell(row=row, column=11, value=it.get("unit") or "")
            _shixun_border(ws, row, row, 1, 11)
            row += 1
            ws.cell(row=row, column=1, value=(it.get("quota_code") or "") or "—")
            ws.cell(row=row, column=2, value=(it.get("quota_name") or "") or "（未匹配定额）")
            ws.cell(row=row, column=3, value=it.get("unit") or "")
            ws.cell(row=row, column=4, value=1)
            qty = float(it.get("quantity") or 1) or 1
            ws.cell(row=row, column=5, value=round(float(it.get("total") or 0) / qty, 2))
            ws.cell(row=row, column=6, value=round(float(it.get("total") or 0), 2))
            _shixun_border(ws, row, row, 1, 11)
            row += 2
    for idx, w in enumerate([12, 26, 12, 10, 12, 14, 12, 22, 12, 14, 12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # ================= 表-11 总价措施项目清单与计价表 =================
    ws = wb.create_sheet("表-11 总价措施项目清单与计价表")
    _shixun_title(ws, 2, 10, "总价措施项目清单与计价表", size=14)
    ws["A3"] = f"工程名称：{project_name}"; ws.merge_cells("A3:C3")
    ws["E3"] = "第  1  页  共  1  页"
    _shixun_header(ws, 4, ["序号", "项目编码", "项目名称", "计算基础", "费率(%)", "金额(元)", "调整后金额(元)", "备注"])
    measures = [
        (9, "011707001001", "安全文明施工", "分部分项人工费+分部分项机械费", 0, 0, "演示默认"),
        (10, "011707001002", "其中：临时设施费", "分部分项人工费+分部分项机械费", 0, 0, ""),
        (13, "011707002001", "夜间施工增加费", "按实际发生计取", 0, 0, "不发生不计取"),
    ]
    for i, (seq, code, name, base, rate, val, note) in enumerate(measures, 5):
        ws.cell(row=i, column=1, value=seq); ws.cell(row=i, column=2, value=code)
        ws.cell(row=i, column=3, value=name); ws.cell(row=i, column=4, value=base)
        ws.cell(row=i, column=5, value=rate); ws.cell(row=i, column=6, value=val)
        ws.cell(row=i, column=7, value=val); ws.cell(row=i, column=8, value=note)
    _shixun_border(ws, 4, 4 + len(measures), 1, 8)
    for idx, w in enumerate([8, 16, 30, 36, 10, 12, 14, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # ================= 表-12 其他项目清单与计价汇总表 =================
    ws = wb.create_sheet("表-12 其他项目清单与计价汇总表")
    _shixun_title(ws, 2, 6, "其他项目清单与计价汇总表", size=14)
    ws["A3"] = f"工程名称：{project_name}"; ws.merge_cells("A3:C3")
    ws["E3"] = "第  1  页  共  1  页"
    _shixun_header(ws, 4, ["序号", "项目名称", "金额(元)", "结算金额(元)", "备注"])
    others = [
        ("1", "暂列金额", 0, 0, "明细详见表-12-1"),
        ("2", "计税暂估价", 0, 0, ""),
        ("2.1", "材料暂估价", 0, 0, "明细详见表-12-2"),
        ("2.2", "专业工程暂估价", 0, 0, "明细详见表-12-3"),
        ("3", "计日工", 0, 0, "明细详见表-12-4"),
        ("4", "总承包服务费", 0, 0, "明细详见表-12-5"),
        ("", "合计", 0, 0, ""),
    ]
    for i, (seq, name, val, settle, note) in enumerate(others, 5):
        ws.cell(row=i, column=1, value=seq); ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=val); ws.cell(row=i, column=4, value=settle)
        ws.cell(row=i, column=5, value=note)
    _shixun_border(ws, 4, 4 + len(others), 1, 5)
    for idx, w in enumerate([8, 26, 14, 14, 26], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # ============ 表-12-1 ~ 表-12-5 明细与表-13（统一生成） ============
    def _blank_detail(name: str, headers: list[str], widths: list[int], rows: list[list]) -> None:
        w_ = wb.create_sheet(name)
        _shixun_title(w_, 2, len(headers), name, size=14)
        w_["A3"] = f"工程名称：{project_name}"; w_.merge_cells(f"A3:{openpyxl.utils.get_column_letter(len(headers))}3")
        _shixun_header(w_, 4, headers)
        for i, r in enumerate(rows, 5):
            for j, v in enumerate(r, 1):
                w_.cell(row=i, column=j, value=v)
        _shixun_border(w_, 4, 4 + len(rows), 1, len(headers))
        for idx, w in enumerate(widths, 1):
            w_.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    _blank_detail(
        "表-12-1 暂列金额明细表",
        ["序号", "项目名称", "计量单位", "暂定金额(元)", "备注"],
        [8, 30, 12, 16, 22],
        [[1, "工程造价上涨", "项", 0, ""], [2, "建筑工程", "项", 0, ""], [3, "装饰工程", "项", 0, ""]],
    )
    _blank_detail(
        "表-12-3 专业工程暂估价及结算价表",
        ["序号", "工程名称", "工程内容", "暂估金额（元）", "结算金额(元)", "差额±(元)", "备注"],
        [8, 22, 30, 16, 16, 16, 24],
        [[1, "地基处理", "本项目未涉及，按 0 填报", 0, 0, 0, ""]],
    )
    _blank_detail(
        "表-12-4 计日工表",
        ["编号", "项目名称", "单位", "暂定数量", "实际数量", "综合单价(元)", "合价(暂定)", "备注"],
        [8, 24, 10, 12, 12, 14, 14, 22],
        [
            ["1", "人工", "", "", "", "", "", ""],
            ["1.1", "土建综合工日", "工日", 0, None, 0, 0, "本项目未发生"],
            ["2", "材料", "", "", "", "", "", ""],
            ["2.1", "材料小计", "", 0, None, 0, 0, "本项目未发生"],
        ],
    )
    _blank_detail(
        "表-12-5 总承包服务费计价表",
        ["序号", "项目名称", "项目价值(元)", "服务内容", "金额(元)", "备注"],
        [8, 26, 16, 40, 14, 22],
        [
            [1, "专业发包工程管理费", 0, "施工质量、进度管理；竣工资料管理", 0, "本项目未发生"],
            [2, "甲供材料设备保管费", 0, "", 0, "本项目未发生"],
        ],
    )
    _blank_detail(
        "表-13 规费税金项目清单与计价表",
        ["序号", "项目名称", "计算基础", "金额（元）", "备注"],
        [8, 30, 44, 14, 24],
        [
            ["1", "规费", "社会保险费+住房公积金", 0, ""],
            ["1.1", "社会保险费", "养老+失业+医疗+工伤", 0, ""],
            ["1.11", "养老保险费", "分部分项人工费+分部分项机械费", 0, ""],
            ["1.12", "失业保险费", "同上", 0, ""],
            ["1.13", "医疗保险费", "同上", 0, ""],
            ["2", "税金", f"招标控制价合计 × 9%（演示 {tax:,.2f} 元）", tax, "增值税"],
        ],
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_task_excel(task_id: str, task: dict) -> bytes:
    return _build_shixun_workbook_bytes(task_id, task)
    # ===== 以下为旧导出逻辑，保留作对照，不再执行 =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工程量识别"

    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)

    ws.merge_cells("A1:L1")
    ws["A1"] = "图纸解析工程量报表"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A2"] = "任务编号"
    ws["B2"] = task_id
    ws["A3"] = "生成时间"
    ws["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A4"] = "解析摘要"
    ws["B4"] = task.get("summary", "")
    ws["B4"].alignment = wrap

    headers = [
        "编号", "构件类型", "图元数", "主规格", "单位", "工程量",
        "线长(m)", "面积(m²)", "置信度", "材料", "来源图层", "计算说明",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, item in enumerate(task.get("components", []), 7):
        ws.cell(row=row_idx, column=1, value=item.get("id", ""))
        ws.cell(row=row_idx, column=2, value=item.get("type", ""))
        ws.cell(row=row_idx, column=3, value=item.get("count", 0))
        ws.cell(row=row_idx, column=4, value=item.get("spec", ""))
        ws.cell(row=row_idx, column=5, value=item.get("unit", ""))
        ws.cell(row=row_idx, column=6, value=item.get("quantity_estimate", 0))
        ws.cell(row=row_idx, column=7, value=item.get("length_m", 0))
        ws.cell(row=row_idx, column=8, value=item.get("area_m2", 0))
        ws.cell(row=row_idx, column=9, value=item.get("confidence", 0))
        ws.cell(row=row_idx, column=10, value=item.get("material", ""))
        ws.cell(row=row_idx, column=11, value="、".join(item.get("layers", [])))
        ws.cell(row=row_idx, column=12, value=item.get("calc_note", ""))
        ws.cell(row=row_idx, column=12).alignment = wrap

    widths = [10, 14, 10, 16, 8, 12, 12, 12, 10, 12, 24, 42]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("清单建议")
    suggestion_headers = ["源构件", "清单编码", "清单名称", "单位", "建议工程量", "项目特征", "置信度", "材料", "构件数"]
    for col, header in enumerate(suggestion_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row_idx, item in enumerate(task.get("boq_suggestions", []), 2):
        ws2.cell(row=row_idx, column=1, value=item.get("source_component_id", ""))
        ws2.cell(row=row_idx, column=2, value=item.get("suggested_code", ""))
        ws2.cell(row=row_idx, column=3, value=item.get("suggested_name", ""))
        ws2.cell(row=row_idx, column=4, value=item.get("suggested_unit", ""))
        ws2.cell(row=row_idx, column=5, value=item.get("suggested_quantity", 0))
        ws2.cell(row=row_idx, column=6, value=item.get("characteristics", ""))
        ws2.cell(row=row_idx, column=7, value=item.get("confidence", 0))
        ws2.cell(row=row_idx, column=8, value=item.get("material", ""))
        ws2.cell(row=row_idx, column=9, value=item.get("component_count", 0))
        ws2.cell(row=row_idx, column=6).alignment = wrap
    for idx, width in enumerate([12, 14, 28, 8, 12, 44, 10, 12, 10], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    ws3 = wb.create_sheet("图层证据")
    layer_headers = ["图层", "图元数", "归类", "图元类型"]
    for col, header in enumerate(layer_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row_idx, item in enumerate(task.get("layer_summary", []), 2):
        entity_types = item.get("entity_types", {})
        ws3.cell(row=row_idx, column=1, value=item.get("layer", ""))
        ws3.cell(row=row_idx, column=2, value=item.get("count", 0))
        ws3.cell(row=row_idx, column=3, value=item.get("classified_as", ""))
        ws3.cell(
            row=row_idx,
            column=4,
            value="、".join(f"{k}:{v}" for k, v in entity_types.items()),
        )
    for idx, width in enumerate([26, 10, 14, 38], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    ws4 = wb.create_sheet("诊断信息")
    ws4.cell(row=1, column=1, value="诊断信息").font = header_font
    for row_idx, item in enumerate(task.get("diagnostics", []), 2):
        ws4.cell(row=row_idx, column=1, value=item)
        ws4.cell(row=row_idx, column=1).alignment = wrap
    ws4.column_dimensions["A"].width = 80

    valuation = task.get("valuation") or {}
    if valuation:
        ws5 = wb.create_sheet("自动计价")
        ws5.merge_cells("A1:J1")
        ws5["A1"] = "图纸解析自动计价结果"
        ws5["A1"].font = title_font
        ws5["A1"].alignment = center
        ws5["A2"] = "项目编号"
        ws5["B2"] = valuation.get("project_id") or ""
        ws5["C2"] = "项目名称"
        ws5["D2"] = valuation.get("project_name") or ""
        ws5["A3"] = "清单项"
        ws5["B3"] = valuation.get("boq_items_created", 0)
        ws5["C3"] = "已匹配定额"
        ws5["D3"] = valuation.get("matched", 0)
        ws5["E3"] = "未匹配"
        ws5["F3"] = valuation.get("skipped", 0)
        ws5["G3"] = "总价"
        ws5["H3"] = valuation.get("grand_total", 0)
        if valuation.get("error"):
            ws5["A4"] = "提示"
            ws5["B4"] = valuation.get("error")

        valuation_headers = [
            "清单ID", "清单编码", "清单名称", "单位", "工程量",
            "定额编码", "定额名称", "匹配度", "状态", "合价",
        ]
        for col, header in enumerate(valuation_headers, 1):
            cell = ws5.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for row_idx, item in enumerate(valuation.get("items", []), 7):
            ws5.cell(row=row_idx, column=1, value=item.get("boq_item_id", ""))
            ws5.cell(row=row_idx, column=2, value=item.get("code", ""))
            ws5.cell(row=row_idx, column=3, value=item.get("name", ""))
            ws5.cell(row=row_idx, column=4, value=item.get("unit", ""))
            ws5.cell(row=row_idx, column=5, value=item.get("quantity", 0))
            ws5.cell(row=row_idx, column=6, value=item.get("quota_code", ""))
            ws5.cell(row=row_idx, column=7, value=item.get("quota_name", ""))
            ws5.cell(row=row_idx, column=8, value=item.get("match_confidence", 0))
            ws5.cell(row=row_idx, column=9, value=item.get("status", ""))
            ws5.cell(row=row_idx, column=10, value=item.get("total", 0))
            ws5.cell(row=row_idx, column=3).alignment = wrap
            ws5.cell(row=row_idx, column=7).alignment = wrap
        for idx, width in enumerate([10, 14, 30, 8, 12, 16, 32, 10, 10, 14], 1):
            ws5.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _run_recognition(
    task_id: str,
    file_bytes: bytes,
    filename: str,
    content_type: str,
    project_context: str,
    file_hash: str = "",
) -> None:
    try:
        is_cad = (
            filename.lower().endswith((".dxf", ".dwg"))
            or "dxf" in content_type.lower()
            or "dwg" in content_type.lower()
        )

        if is_cad:
            _store_result(task_id, {"progress": "正在解析 CAD 图层和图元..."})

            # 快速看图几何提前到管线最前（约十余秒），前端 WebGL 秒开看图；
            # 高清原图在构件分析之后渲染（大图数分钟），完成后推送替换
            _phase_pct = {
                "转换": 8, "读取 CAD": 12, "图层和模型空间": 18,
                "收集图纸图元": 22, "快速看图几何": 25, "快速看图已就绪": 28,
                "分类图元": 34, "标记构件": 45, "汇总构件": 52,
                "高清原图": 55,
                "高清预览·准备": 60, "高清预览·绘制": 78, "高清预览·输出": 90,
                "高清原图已生成": 93,
                "生成图纸预览": 90, "生成高清预览": 92, "整理识别结果": 95,
            }
            _prog_state: dict[str, int] = {"last": 18}
            def _cad_progress(payload: dict) -> None:
                payload = dict(payload)
                extra: dict = {}
                # 大体积数据（渲染图/几何）通过进度通道送达：单独存放并打就绪标记，
                # 轻量轮询只传标记，前端看到标记后再按需拉取
                raster = payload.pop("cad_raster", None)
                if raster:
                    with _raster_lock:
                        _RASTER_STORE[task_id] = raster
                    extra["cad_raster_ready"] = True
                geometry = payload.pop("cad_geometry", None)
                if geometry:
                    with _geometry_lock:
                        _GEOMETRY_STORE[task_id] = geometry
                    extra["cad_geometry_ready"] = True
                text = payload.get("progress", "")
                explicit_pct = payload.pop("progress_percent", None)
                if explicit_pct is not None:
                    # 渲染心跳按图元数插值，本身单调，直接采信；
                    # 否则会与下方 last+4 爬升叠加，把百分比顶到 93 后锁死
                    try:
                        pct = max(int(explicit_pct), _prog_state["last"])
                    except (TypeError, ValueError):
                        pct = _prog_state["last"]
                else:
                    pct = 30
                    for key, val in _phase_pct.items():
                        if key in text:
                            pct = val
                            break
                    else:
                        pct = min(_prog_state["last"] + 4, 93)
                    # 单调递增防护：关键词映射值可能低于先前平滑累积值（进度回跳），取较大值
                    pct = max(pct, _prog_state["last"])
                _prog_state["last"] = pct
                _store_result(task_id, {**payload, **extra, "progress_percent": pct})

            raw = analyze_dxf_bytes(file_bytes, filename, progress_callback=_cad_progress)
            components = [_dump_model(ComponentOut(**item)) for item in raw.get("components", [])]
            suggestions = [_dump_model(BoqSuggestionOut(**item)) for item in raw.get("boq_suggestions", [])]
            has_error = bool(raw.get("error"))
            _store_result(task_id, {
                "status": "error" if has_error else "done",
                "drawing_type": raw.get("drawing_type", ""),
                "summary": raw.get("summary", ""),
                "components": components,
                "boq_suggestions": suggestions,
                "diagnostics": raw.get("diagnostics", []),
                "layer_summary": raw.get("layer_summary", []),
                "disciplines": raw.get("disciplines", []),
                "quality_score": raw.get("quality_score"),
                "preview_svg": raw.get("preview_svg", ""),
                "preview_svg_hd": raw.get("preview_svg_hd", ""),
                "cad_geometry": raw.get("cad_geometry"),
                "cad_raster": raw.get("cad_raster"),
                "cad_raster_ready": bool(raw.get("cad_raster")),
                "cad_geometry_ready": bool((raw.get("cad_geometry") or {}).get("bbox")),
                "valuation": None,
                "valuation_status": "processing" if suggestions and not has_error else "skipped",
                "valuation_progress": "识别完成，正在准备自动计价..." if suggestions and not has_error else "未生成可计价清单",
                "valuation_progress_percent": 0,
                "valuation_error": None,
                "progress": "解析完成" if not has_error else "解析失败",
                "error": raw.get("error"),
            })
            if suggestions and not has_error:
                if file_hash:
                    _RESULT_CACHE[file_hash] = task_id
                _start_drawing_valuation(task_id)
            return

        _store_result(task_id, {"progress": "正在调用 辅助识别图纸..."})
        result = recognize_drawing(
            image_bytes=file_bytes,
            content_type=content_type,
            project_context=project_context,
        )
        components = [
            _dump_model(ComponentOut(
                id=c.id,
                type=c.type,
                count=c.count,
                spec=c.spec,
                confidence=c.confidence,
                material=c.material,
                unit=c.unit,
                quantity_estimate=c.quantity_estimate,
                calc_note="辅助 视觉识别结果，工程量由模型估算",
            ))
            for c in (result.components or [])
        ]
        suggestions = [
            _dump_model(BoqSuggestionOut(**item))
            for item in (components_to_boq_suggestions(result.components) if result.components else [])
        ]
        has_error = bool(result.error)
        _store_result(task_id, {
            "status": "error" if has_error else "done",
            "drawing_type": result.drawing_type,
            "summary": result.summary,
            "components": components,
            "boq_suggestions": suggestions,
            "diagnostics": [],
            "layer_summary": [],
            "preview_svg": "",
            "preview_svg_hd": "",
            "valuation": None,
            "valuation_status": "processing" if suggestions and not has_error else "skipped",
            "valuation_progress": "识别完成，正在准备自动计价..." if suggestions and not has_error else "未生成可计价清单",
            "valuation_progress_percent": 0,
            "valuation_error": None,
            "progress": "解析完成" if not has_error else "解析失败",
            "error": result.error,
        })
        if suggestions and not has_error:
            _start_drawing_valuation(task_id)
    except Exception as exc:
        logger.exception("Drawing recognition failed task_id=%s: %s", task_id, exc)
        _store_result(task_id, {
            "status": "error",
            "preview_svg": "",
            "preview_svg_hd": "",
            "valuation": None,
            "valuation_status": "error",
            "valuation_progress": "识别任务异常，自动计价未执行",
            "valuation_progress_percent": 0,
            "valuation_error": f"识别任务异常: {exc}",
            "progress": "识别任务异常",
            "error": f"识别任务异常: {exc}",
        })


@router.post("", summary="上传图纸，返回任务 ID")
async def upload_drawing(
    file: UploadFile = File(..., description="图纸文件 (PNG/JPG/PDF/DXF/DWG)"),
    project_context: str = Query("", description="可选：项目背景描述，提升识别精度"),
):
    task_id = str(uuid.uuid4())
    try:
        file_bytes = await _read_upload_bytes_limited(file)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Drawing upload failed before task creation: %s", exc)
        raise HTTPException(status_code=400, detail=f"读取图纸文件失败: {exc}") from exc
    filename = file.filename or "drawing"
    content_type = file.content_type or "application/octet-stream"

    # 同一张图纸重复上传：直接复用上次解析结果（原地重传验证场景秒出）
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    cached_task_id = _RESULT_CACHE.get(file_hash)
    if cached_task_id:
        cached = _get_task(cached_task_id)
        if cached and cached.get("status") == "done" and cached.get("components"):
            logger.info("图纸内容与任务 %s 相同，直接复用解析结果", cached_task_id)
            return {"taskId": cached_task_id, "cached": True}
        _RESULT_CACHE.pop(file_hash, None)

    # 原始图纸落盘（供"用 CAD 快速看图打开"：本机调起看图软件打开原文件）
    try:
        _DRAWING_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        safe_name = (sanitize_filename(filename) or "drawing").replace("\x00", "")
        save_path = _DRAWING_CACHE_DIR / f"{task_id}_{safe_name}"
        save_path.write_bytes(file_bytes)
        logger.info("原始图纸已落盘：%s", save_path)
    except Exception as exc:
        logger.warning("原始图纸落盘失败（不影响解析）：%s", exc)

    with _tasks_lock:
        now = datetime.now(timezone.utc).isoformat()
        _tasks[task_id] = {
            "status": "processing",
            "drawing_type": "",
            "summary": "",
            "components": [],
            "boq_suggestions": [],
            "diagnostics": [],
            "layer_summary": [],
            "preview_svg": "",
            "preview_svg_hd": "",
            "valuation": None,
            "valuation_status": "idle",
            "valuation_progress": "",
            "valuation_progress_percent": 0,
            "valuation_error": None,
            "progress": "正在读取图纸文件...",
            "created_at": now,
            "updated_at": now,
            "error": None,
        }
        initial_task = dict(_tasks[task_id])
    save_background_task(task_id, _TASK_TYPE, initial_task)

    thread = threading.Thread(
        target=_run_recognition,
        args=(task_id, file_bytes, filename, content_type, project_context, file_hash),
        daemon=True,
    )
    thread.start()

    return {"taskId": task_id}


@router.post("/convert/dxf-to-dwg", summary="将 DXF 转换为 DWG")
async def convert_dxf_to_dwg(
    file: UploadFile = File(..., description="DXF 图纸文件"),
):
    filename = file.filename or "drawing.dxf"
    if not filename.lower().endswith(".dxf"):
        raise HTTPException(status_code=400, detail="请上传 DXF 文件")

    file_bytes = await _read_upload_bytes_limited(file)
    # 外部转换器子进程最长可跑数分钟，放线程池执行，避免阻塞事件循环
    result = await run_in_threadpool(convert_dxf_to_dwg_bytes, file_bytes, filename)
    if result.error or result.dwg_bytes is None:
        raise HTTPException(
            status_code=400,
            detail={
                "error": result.error or "dxf_to_dwg_failed",
                "diagnostics": result.diagnostics,
            },
        )

    raw_stem = filename.rsplit(".", 1)[0]
    safe_stem = sanitize_filename(raw_stem) or "drawing"
    out_name = f"{safe_stem}.dwg"
    return StreamingResponse(
        io.BytesIO(result.dwg_bytes),
        media_type="application/acad",
        headers={"Content-Disposition": build_attachment_disposition(out_name)},
    )


@router.get("/convert/status", response_model=ConverterStatusResponse, summary="查询 CAD 转换器状态")
async def get_cad_converter_status():
    # 状态探测含磁盘扫描，放线程池执行
    status = await run_in_threadpool(get_converter_status)
    return ConverterStatusResponse(**status)


@router.post("/{task_id}/embed-cad", summary="在网页预览区嵌入 CAD 快速看图内核")
async def embed_cad_start(task_id: str, rect: dict):
    """调起 CAD 快速看图并以无边框置顶窗口贴合网页预览区。

    前端持续把预览区的屏幕坐标（物理像素）上报到 /embed-cad/rect，
    查看器进程轮询坐标文件实时跟随（浏览器移动/缩放/滚动），
    效果等同把 CAD 看图窗口"嵌"进网页预览区。
    """
    try:
        _DRAWING_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        matches = sorted(_DRAWING_CACHE_DIR.glob(f"{task_id}_*"), key=lambda p: p.stat().st_mtime)
    except Exception:
        matches = []
    if not matches:
        raise HTTPException(status_code=404, detail="未找到该图纸的本地原始文件，请重新上传图纸后再试")

    exe = _find_cad_reader_exe()
    if not exe:
        return {
            "opened": False,
            "message": "未检测到本机 CAD 快速看图（D:\\CADReader\\CADReader.exe）或同类看图软件，请先安装 CAD 快速看图",
        }

    target = str(matches[0])
    # routes(0) → api(1) → app(2) → backend(3)
    viewer_script = Path(__file__).resolve().parents[3] / "tools" / "cad_viewer.py"
    if not viewer_script.is_file():
        return {"opened": False, "message": "查看器脚本缺失（tools/cad_viewer.py）"}

    # 同一时间只保留一个嵌入实例：启动新的先停旧的
    _embed_stop()

    # 会话 id 进文件名：旧查看器轮询的是旧文件，删除旧文件只会让它驻留退出，
    # 不会误读新会话坐标；新查看器只认自己的文件
    session_id = uuid.uuid4().hex[:8]
    rect_file = Path(tempfile.gettempdir()) / f"zjcost_embed_rect_{task_id}_{session_id}.json"
    state_file = rect_file.with_suffix(".state.json")
    try:
        state_file.write_text(json.dumps({"state": "starting"}), encoding="utf-8")
        rect_file.write_text(
            json.dumps({"x": rect.get("x", 0), "y": rect.get("y", 0),
                        "w": rect.get("w", 0), "h": rect.get("h", 0),
                        "visible": bool(rect.get("visible", True)),
                        "pf": bool(rect.get("pf", True))}, ensure_ascii=False),
            encoding="utf-8",
        )
        proc = subprocess.Popen(
            [sys.executable, str(viewer_script), "--embed", str(rect_file), target],
            close_fds=True,
            creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
            cwd=str(viewer_script.parent),
        )
    except Exception as exc:
        logger.exception("调起嵌入式图纸查看器失败：%s", target)
        raise HTTPException(status_code=500, detail=f"调起嵌入式图纸查看器失败: {exc}") from exc
    with _embed_lock:
        _EMBED_STATE.update(task_id=task_id, rect_file=rect_file, state_file=state_file, proc=proc)
    logger.info("已在预览区嵌入 CAD 快速看图：%s", target)

    # 等待查看器把状态写到 attached 再返回成功：attach 表示 CAD 已挂到宿主，
    # 坐标流此时开始灌入；若继续等 ready，而 ready 又依赖前端坐标流，
    # 会形成"前端等 opened → 查看器等 rect"的死锁。
    # 启动失败/超时自动回退内置渲染，不再让前端一直停在"CAD 内核预览启动中…"
    deadline = time.time() + EMBED_START_TIMEOUT
    while time.time() < deadline:
        payload = None
        try:
            payload = json.loads(state_file.read_text(encoding="utf-8"))
        except Exception:
            pass
        state = (payload or {}).get("state", "starting")
        if state in ("attached", "ready"):
            logger.info("CAD 嵌入查看器就绪：%s", target)
            return {"opened": True, "app": "cad_viewer_embed", "file": Path(target).name}
        if state in ("failed", "exited", "error"):
            _embed_stop()
            return {
                "opened": False,
                "message": (payload or {}).get("message") or "CAD 内核预览启动失败，已切换内置渲染",
            }
        if proc.poll() is not None:
            _embed_stop()
            return {"opened": False, "message": "CAD 内核预览进程已退出，已切换内置渲染"}
        await asyncio.sleep(0.25)

    _embed_stop()
    return {"opened": False, "message": "CAD 内核预览启动超时，已切换内置渲染"}


@router.post("/{task_id}/embed-cad/rect", summary="上报预览区屏幕坐标（嵌入模式跟随）")
async def embed_cad_rect(task_id: str, rect: dict):
    with _embed_lock:
        if _EMBED_STATE["task_id"] != task_id or not _EMBED_STATE["rect_file"]:
            return {"ok": False}
        rect_file = _EMBED_STATE["rect_file"]
    try:
        rect_file.write_text(
            json.dumps({"x": rect.get("x", 0), "y": rect.get("y", 0),
                        "w": rect.get("w", 0), "h": rect.get("h", 0),
                        "visible": bool(rect.get("visible", True)),
                        "pf": bool(rect.get("pf", True))}, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception:
        return {"ok": False}
    return {"ok": True}


@router.post("/{task_id}/embed-cad/stop", summary="停止嵌入的 CAD 快速看图")
async def embed_cad_stop(task_id: str):
    with _embed_lock:
        if _EMBED_STATE["task_id"] != task_id:
            return {"ok": False}
    _embed_stop()
    return {"ok": True}


@router.get("/{task_id}/embed-cad/status", summary="查询嵌入的 CAD 快速看图状态")
async def embed_cad_status(task_id: str):
    """查看器启动/退出状态查询：前端据此在启动失败或运行中途退出时回退内置渲染。"""
    with _embed_lock:
        if _EMBED_STATE["task_id"] != task_id:
            return {"state": "not_found", "message": ""}
        state_file = _EMBED_STATE["state_file"]
        proc = _EMBED_STATE["proc"]
    message = ""
    state = "starting"
    if state_file and state_file.is_file():
        try:
            payload = json.loads(state_file.read_text(encoding="utf-8"))
            state = payload.get("state", "starting")
            message = payload.get("message", "") or ""
        except Exception:
            pass
    if proc is not None and proc.poll() is not None and state not in ("exited", "failed", "error"):
        state = "exited"
        message = "CAD 内核预览进程已退出"
    return {"state": state, "message": message}


def _embed_stop() -> None:
    """停止当前嵌入实例：删除坐标文件（查看器轮询到消失即自行退出）。"""
    with _embed_lock:
        rect_file = _EMBED_STATE["rect_file"]
        state_file = _EMBED_STATE["state_file"]
        _EMBED_STATE.update(task_id="", rect_file=None, state_file=None, proc=None)
    if rect_file:
        try:
            rect_file.unlink(missing_ok=True)
        except Exception:
            pass
    if state_file:
        try:
            state_file.unlink(missing_ok=True)
        except Exception:
            pass


@router.get("/{task_id}/geometry", summary="获取快速看图几何数据（解析中可提前拉取）")
async def get_drawing_geometry(task_id: str):
    with _geometry_lock:
        geometry = _GEOMETRY_STORE.get(task_id)
    if geometry is None:
        task = _get_task(task_id)
        geometry = (task or {}).get("cad_geometry")
    if not geometry or not geometry.get("bbox"):
        raise HTTPException(status_code=404, detail="快速看图几何尚未生成")
    return {"cad_geometry": geometry}


@router.get("/{task_id}/raster", summary="获取 CAD 原图高清渲染图（解析中可提前拉取）")
async def get_drawing_raster(task_id: str):
    with _raster_lock:
        raster = _RASTER_STORE.get(task_id)
    if raster is None:
        task = _get_task(task_id)
        raster = (task or {}).get("cad_raster")
    if not raster or not raster.get("data_url"):
        raise HTTPException(status_code=404, detail="渲染图尚未生成")
    return {"cad_raster": raster}


@router.get("/{task_id}", response_model=TaskStatusResponse, summary="查询识别结果")
async def get_recognition_result(task_id: str, include_svg: bool = Query(True)):
    task = _get_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail=f"任务不存在: {task_id}")

    # 大体积数据一律走专用按需端点（/raster、/geometry，前端见就绪标记后拉取）：
    # 渲染图 base64 可达数十 MB，内联在轮询/完整结果里会把页面拖卡甚至超时
    task = {**task, "cad_geometry": None, "cad_raster": None}
    if not include_svg or task.get("status") == "processing":
        # 预览 SVG 可达数 MB：解析阶段前端用不到，
        # 轮询反复传输会把页面拖卡，仅在终态且显式请求时返回
        task = {**task, "preview_svg": "", "preview_svg_hd": ""}

    return TaskStatusResponse(taskId=task_id, **task)


@router.get("/{task_id}/export", summary="导出图纸解析工程量 Excel")
async def export_recognition_result(task_id: str):
    task = _get_task(task_id) or {}
    if not task:
        raise HTTPException(status_code=404, detail=f"任务不存在: {task_id}")
    if task.get("status") == "processing":
        raise HTTPException(status_code=409, detail="图纸仍在解析中，请稍后再导出")

    # openpyxl 生成 Excel 是重同步操作，放线程池执行
    file_bytes = await run_in_threadpool(_export_task_excel, task_id, task)
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_safe_report_filename(task_id, "xlsx")}"'},
    )
