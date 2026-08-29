"""BOQ import template download endpoint.

Generates a standard Excel template for BOQ list import,
reducing format mismatch errors.
"""

import io
import openpyxl
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.session import get_db

router = APIRouter(prefix="/templates", tags=["templates"])


@router.get("/boq-import.xlsx")
def download_boq_template():
    """Download a standard BOQ import Excel template.
    
    Columns match the fields expected by the BOQ import API.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "清单导入模板"

    # Header row with descriptions
    headers = [
        ("编码", "清单项编码，如 010101001"),
        ("名称", "清单项名称，如 挖一般土方"),
        ("项目特征", "项目特征描述"),
        ("计量单位", "如 m³、m²、t"),
        ("工程量", "数值，如 100.5"),
        ("专业", "如 土建、安装、装饰"),
        ("分部", "如 A.1 土石方工程"),
        ("备注", "补充说明"),
    ]

    for col_idx, (header, note) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True, size=11)
        cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        # Example row
        example_cell = ws.cell(row=2, column=col_idx, value=note)
        example_cell.font = openpyxl.styles.Font(color="808080", italic=True, size=10)

    # Sample data row
    sample = ["010101001001", "挖一般土方", "三类土，深1.5m内", "m³", "100", "土建", "A.1 土石方工程", ""]
    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=3, column=col_idx, value=val)

    # Column widths
    widths = [18, 20, 30, 10, 10, 8, 20, 15]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # Data validation for unit column
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"m,m²,m³,t,kg,个,项,套,台,组,根,块,樘,扇,座,处"', allow_blank=True)
    dv.error = "请使用标准计量单位"
    dv.errorTitle = "单位格式"
    ws.add_data_validation(dv)
    dv.add(f"D3:D1000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=boq_import_template.xlsx"},
    )


@router.get("/quota-import.xlsx")
def download_quota_template():
    """Download a standard quota library import template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "定额导入模板"

    headers = [
        ("定额号", "如 1-1"),
        ("定额名称", "如 人工挖土方"),
        ("计量单位", "如 100m³"),
        ("人工费", "元"),
        ("材料费", "元"),
        ("机械费", "元"),
        ("基价", "元（可空，自动计算）"),
        ("类别", "如 土建、安装"),
    ]

    for col_idx, (header, note) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
        cell.fill = openpyxl.styles.PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col_idx, value=note).font = openpyxl.styles.Font(color="808080", italic=True, size=10)

    sample = ["1-1", "人工挖土方 深度1.5m内", "100m³", "1200", "0", "0", "1200", "土建"]
    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=3, column=col_idx, value=val)

    widths = [12, 25, 10, 10, 10, 10, 10, 8]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quota_import_template.xlsx"},
    )


@router.get("/material-price-import.xlsx")
def download_material_price_template():
    """Download a material price import template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "材料价导入模板"

    headers = [
        ("材料编码", "如 C01001"),
        ("材料名称", "如 普通硅酸盐水泥 P.O 42.5"),
        ("规格型号", "如 袋装 50kg"),
        ("单位", "如 t"),
        ("单价(元)", "含税市场价"),
        ("来源", "如 信息价、市场询价、合同价"),
        ("日期", "如 2026-06-01"),
    ]

    for col_idx, (header, note) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
        cell.fill = openpyxl.styles.PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col_idx, value=note).font = openpyxl.styles.Font(color="808080", italic=True, size=10)

    sample = ["C01001", "普通硅酸盐水泥 P.O 42.5", "袋装 50kg", "t", "580", "信息价", "2026-06-01"]
    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=3, column=col_idx, value=val)

    widths = [12, 25, 15, 8, 10, 12, 12]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=material_price_import_template.xlsx"},
    )
